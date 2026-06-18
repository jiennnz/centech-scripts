from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from financial.sales_export_comparison.rules import OrgRule, SideInputConfig


@dataclass(frozen=True)
class WorkbookFillConfig:
    workbook_path: Path
    org_rule: OrgRule
    centech_path: Path
    start_date: date
    end_date: date
    client_path: Path | None = None
    client_side_config: "SideInputConfig | None" = None  # overrides org_rule.client when set
    centech_side_config: "SideInputConfig | None" = None  # overrides org_rule.centech when set


@dataclass(frozen=True)
class WorkbookFillResult:
    centech_rows_written: int
    client_rows_written: int
    ignored_categories: frozenset[str] = frozenset()


def _process_side(
    wb,
    df: pd.DataFrame,
    side_cfg,
    category_rows: dict[str, int],
    *,
    column_offsets: tuple[int, int],
    start_date: date,
    end_date: date,
    sheet_date_format: str,
) -> int:
    written = 0
    dates = _parse_dates(df[side_cfg.date_column], side_cfg.date_parse_format)

    for i in range(len(df)):
        row = df.iloc[i]
        ts = dates.iloc[i]
        if pd.isna(ts):
            continue

        row_date = ts.date() if hasattr(ts, "date") else None
        if row_date is None or row_date < start_date or row_date > end_date:
            continue

        debit = row[side_cfg.debit_column]
        credit = row[side_cfg.credit_column]
        debit = 0 if pd.isna(debit) else float(debit)
        credit = 0 if pd.isna(credit) else float(credit)

        category = str(row[side_cfg.category_column]).strip() if pd.notna(row[side_cfg.category_column]) else ""
        if side_cfg.skip_category and category == side_cfg.skip_category:
            continue
        if side_cfg.skip_zero_debit_credit and debit == 0 and credit == 0:
            continue

        had_prefix = bool(side_cfg.category_strip_prefix and category.startswith(side_cfg.category_strip_prefix))
        if had_prefix:
            category = category[len(side_cfg.category_strip_prefix):]
        category = _apply_category_rewrites(category, side_cfg.category_rewrites)
        category = _apply_category_starts_with(category, side_cfg.category_starts_with)
        if had_prefix and side_cfg.category_strip_prefix_fallback and category not in category_rows:
            category = side_cfg.category_strip_prefix_fallback

        if side_cfg.online_credit_card and side_cfg.memo_column:
            memo_val = row[side_cfg.memo_column] if side_cfg.memo_column in row.index else ""
            memo = str(memo_val).strip() if pd.notna(memo_val) else ""
            occ_result = _apply_online_credit_card(category, memo, side_cfg.online_credit_card)
            if occ_result is None:
                continue
            category = occ_result

        debit, credit = _maybe_round_for_category(category, debit, credit, side_cfg)

        store = _parse_store(row[side_cfg.store_column], side_cfg)
        sheet_name = ts.date().strftime(sheet_date_format) if hasattr(ts, "date") else None
        if not sheet_name or sheet_name not in wb.sheetnames:
            continue

        sheet = wb[sheet_name]
        store_columns = _build_store_columns(sheet, column_offsets)
        if _put_data(sheet, store_columns, store, category, debit, credit, category_rows):
            written += 1

    return written


def _col_to_num(col: str) -> int:
    n = 0
    for ch in col.upper():
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def _num_to_col(n: int) -> str:
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(r + ord("A")) + s
    return s


def _next_column(col: str, step: int = 1) -> str:
    return _num_to_col(_col_to_num(col) + step)


def _build_store_columns(sheet: Worksheet, column_offset: tuple[int, int]) -> dict[str, tuple[str, str]]:
    store_cols: dict[str, tuple[str, str]] = {}
    for row in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            val = cell.value
            if not val or not isinstance(val, str) or "Store" not in val:
                continue
            parts = val.split()
            if len(parts) >= 2 and parts[1].isdigit():
                store_num = parts[1]
                col_letter = cell.column_letter
                debit_col = _next_column(col_letter, column_offset[0])
                credit_col = _next_column(col_letter, column_offset[1])
                store_cols[store_num] = (debit_col, credit_col)
    return store_cols


def _put_data(
    sheet: Worksheet,
    store_columns: dict[str, tuple[str, str]],
    store: str,
    category: str,
    debit: float,
    credit: float,
    category_rows: dict[str, int],
) -> bool:
    if store not in store_columns:
        return False
    col_d, col_e = store_columns[store]
    row_num = category_rows.get(category)
    if not row_num:
        return False

    debit_cell = sheet[f"{col_d}{row_num}"]
    credit_cell = sheet[f"{col_e}{row_num}"]

    existing_debit = debit_cell.value if debit_cell.value is not None else 0
    existing_credit = credit_cell.value if credit_cell.value is not None else 0
    if isinstance(existing_debit, str) and existing_debit.startswith("="):
        return False
    if isinstance(existing_credit, str) and existing_credit.startswith("="):
        return False

    try:
        existing_debit = float(existing_debit)
    except (TypeError, ValueError):
        existing_debit = 0.0
    try:
        existing_credit = float(existing_credit)
    except (TypeError, ValueError):
        existing_credit = 0.0

    new_debit = existing_debit + float(debit or 0)
    new_credit = existing_credit + float(credit or 0)

    debit_cell.value = None if new_debit == 0 else new_debit
    credit_cell.value = None if new_credit == 0 else new_credit
    return True


def _parse_store(raw: str, side_cfg) -> str:
    raw = str(raw or "").strip()
    if side_cfg.store_parse == "first_token":
        return raw.split()[0].split("-")[0].strip()
    return raw


def _apply_category_rewrites(category: str, rewrites: dict[str, str]) -> str:
    cat = str(category or "").strip()
    if cat in rewrites:
        return rewrites[cat]
    return cat


def _apply_category_starts_with(category: str, starts_with_map: dict[str, str]) -> str:
    cat = str(category or "").strip()
    lower = cat.lower()
    for prefix, target in starts_with_map.items():
        if lower.startswith(prefix.lower()):
            return target
    return cat


def _apply_online_credit_card(category: str, memo: str, occ: dict) -> str | None:
    """Returns new category, or None to skip row."""
    cat_eq = str(occ.get("category_equals", "Online Credit Card"))
    if category.strip().lower() != cat_eq.lower():
        return category

    memo_eq = str(occ.get("memo_equals", ""))
    map_to = str(occ.get("map_to", "Online credit card"))
    skip_no = bool(occ.get("skip_if_memo_no_match", True))

    if memo.strip() == memo_eq:
        return map_to
    if skip_no:
        return None
    return category


def _maybe_round_for_category(category: str, debit: float, credit: float, side_cfg) -> tuple[float, float]:
    if category not in side_cfg.round_debit_credit_for_categories:
        return debit, credit
    return round(debit, 2), round(credit, 2)


def _normalized_raw_categories(
    df: pd.DataFrame,
    side_cfg,
    *,
    start_date: date,
    end_date: date,
) -> set[str]:
    categories: set[str] = set()
    dates = _parse_dates(df[side_cfg.date_column], side_cfg.date_parse_format)
    for i in range(len(df)):
        ts = dates.iloc[i]
        if pd.isna(ts):
            continue
        row_date = ts.date() if hasattr(ts, "date") else None
        if row_date is None or row_date < start_date or row_date > end_date:
            continue
        debit = df.iloc[i][side_cfg.debit_column]
        credit = df.iloc[i][side_cfg.credit_column]
        debit = 0 if pd.isna(debit) else float(debit)
        credit = 0 if pd.isna(credit) else float(credit)
        if side_cfg.skip_zero_debit_credit and debit == 0 and credit == 0:
            continue
        raw = df.iloc[i][side_cfg.category_column]
        category = str(raw).strip() if pd.notna(raw) else ""
        if not category:
            continue
        if side_cfg.category_strip_prefix and category.startswith(side_cfg.category_strip_prefix):
            category = category[len(side_cfg.category_strip_prefix):]
        categories.add(category.strip())
    return categories


def _client_uses_combined_olo_format(
    df: pd.DataFrame,
    side_cfg,
    *,
    start_date: date,
    end_date: date,
) -> bool:
    categories = _normalized_raw_categories(
        df,
        side_cfg,
        start_date=start_date,
        end_date=end_date,
    )
    if "OLO CC Deposit" not in categories:
        return False
    standalone_online = {
        "Online Credit Card",
        "Online Credit card",
        "Online Gift Card",
    }
    return not bool(categories & standalone_online)


def _read_side_frame(path: Path, side_cfg) -> pd.DataFrame:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        kwargs = dict(side_cfg.read_csv_kwargs)
        return pd.read_csv(path, **kwargs)
    if suffix in {".xlsx", ".xls"}:
        return pd.read_excel(path, engine="openpyxl", **side_cfg.read_excel_kwargs)

    if side_cfg.format == "excel":
        return pd.read_excel(path, engine="openpyxl", **side_cfg.read_excel_kwargs)
    kwargs = dict(side_cfg.read_csv_kwargs)
    return pd.read_csv(path, **kwargs)


def _parse_dates(series: pd.Series, parse_format: str | None) -> pd.Series:
    if parse_format:
        return pd.to_datetime(series, format=parse_format, errors="coerce")
    return pd.to_datetime(series, errors="coerce")


def _auto_select_client_cfg(df: pd.DataFrame, cfg, rule) -> tuple:
    """Return (cfg, df) using whichever client format matches the dataframe's actual columns.

    Checks primary cfg first, then each entry in rule.client_legacy in order.
    Avoids KeyError when the client switches export formats across date ranges.
    """
    cols = set(df.columns)

    def _matches(c) -> bool:
        return {c.date_column, c.store_column, c.category_column,
                c.debit_column, c.credit_column}.issubset(cols)

    if _matches(cfg):
        return cfg, df
    for fallback in getattr(rule, "client_legacy", ()):
        if _matches(fallback):
            return fallback, df
    return cfg, df  # let the original error surface with the real column names


def run(config: WorkbookFillConfig) -> WorkbookFillResult:
    rule = config.org_rule
    category_rows = dict(rule.category_rows)
    if not category_rows:
        raise ValueError("org_rule.category_rows is empty; add category → row map in YAML")

    wb = load_workbook(config.workbook_path)

    centech_cfg = config.centech_side_config or rule.centech
    df_centech = _read_side_frame(config.centech_path, centech_cfg)

    client_cfg = config.client_side_config or rule.client
    client_n = 0
    ignored_categories: set[str] = set()
    if config.client_path is not None:
        df_client = _read_side_frame(config.client_path, client_cfg)
        if config.client_side_config is None:
            client_cfg, df_client = _auto_select_client_cfg(df_client, client_cfg, rule)
        if _client_uses_combined_olo_format(
            df_client,
            client_cfg,
            start_date=config.start_date,
            end_date=config.end_date,
        ):
            ignored_categories.update({"Online Credit card", "Online Gift Card"})
        client_n = _process_side(
            wb,
            df_client,
            client_cfg,
            category_rows,
            column_offsets=(2, 3),
            start_date=config.start_date,
            end_date=config.end_date,
            sheet_date_format=rule.sheet_date_format,
        )

    centech_n = _process_side(
        wb,
        df_centech,
        centech_cfg,
        category_rows,
        column_offsets=(0, 1),
        start_date=config.start_date,
        end_date=config.end_date,
        sheet_date_format=rule.sheet_date_format,
    )

    wb.save(config.workbook_path)
    return WorkbookFillResult(
        centech_rows_written=centech_n,
        client_rows_written=client_n,
        ignored_categories=frozenset(ignored_categories),
    )
