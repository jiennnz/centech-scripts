from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_RUN_ROOT = (
    REPO_ROOT
    / "financial"
    / "sales_export_comparison"
    / "runs"
    / "2026-05-01_2026-05-31"
    / "century"
    / "centech_vs_client"
)


def parse_date(raw: Any) -> date | None:
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    if isinstance(raw, (int, float)):
        return date(1899, 12, 30) + timedelta(days=int(raw))
    if isinstance(raw, str):
        value = raw.strip()
        for fmt in ("%m/%d/%Y", "%Y-%m-%d"):
            try:
                parsed = datetime.strptime(value, fmt)
                return date(parsed.year, parsed.month, parsed.day)
            except ValueError:
                pass
        for fmt in ("%b %d %Y", "%B %d %Y"):
            try:
                parsed = datetime.strptime(f"{value} 2026", fmt)
                return date(parsed.year, parsed.month, parsed.day)
            except ValueError:
                pass
    return None


def parse_cli_date(raw: str) -> date:
    parsed = parse_date(raw)
    if parsed is None:
        raise argparse.ArgumentTypeError(f"Could not parse date: {raw!r}")
    return parsed


def norm_store(raw: Any) -> str | None:
    if raw is None:
        return None
    if isinstance(raw, float) and raw.is_integer():
        return str(int(raw))
    if isinstance(raw, int):
        return str(raw)
    value = str(raw).strip()
    match = re.search(r"(\d+)", value)
    return match.group(1) if match else value


def norm_text(raw: Any) -> str:
    if raw is None:
        return ""
    return re.sub(r"\s+", " ", str(raw).strip())


def norm_account(raw: Any) -> str:
    if raw is None or raw == "":
        return ""
    if isinstance(raw, float) and raw.is_integer():
        return str(int(raw))
    if isinstance(raw, int):
        return str(raw)
    value = str(raw).strip()
    if re.fullmatch(r"\d+\.0", value):
        return value[:-2]
    return value


def number(raw: Any) -> float:
    if raw is None or raw == "":
        return 0.0
    if isinstance(raw, (int, float)):
        return float(raw)
    value = str(raw).strip().replace("$", "").replace(",", "")
    if value.startswith("(") and value.endswith(")"):
        value = "-" + value[1:-1]
    try:
        return float(value)
    except ValueError:
        return 0.0


def header_indexes(row: tuple[Any, ...]) -> dict[str, int]:
    return {norm_text(value): idx for idx, value in enumerate(row) if value is not None}


def load_raa_source_rows(path: Path, start: date, end: date) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = header_indexes(next(ws.iter_rows(min_row=3, max_row=3, values_only=True)))
    required = {"Store", "Date", "Side", "TQSR Amount"}
    missing = sorted(required - set(headers))
    if missing:
        raise ValueError(f"{path}: missing row-3 headers: {missing}")

    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        store = norm_store(row[headers["Store"]])
        row_date = parse_date(row[headers["Date"]])
        side = norm_text(row[headers["Side"]]).title()
        tqsr_amount = number(row[headers["TQSR Amount"]])
        if not store or row_date is None or not side or abs(tqsr_amount) < 0.0000001:
            continue
        if start <= row_date <= end:
            rows.append(
                {
                    "store": store,
                    "date": row_date.isoformat(),
                    "side": side,
                    "tqsr_source": tqsr_amount,
                }
            )
    return rows


def load_comparison_raa_values(path: Path, start: date, end: date) -> dict[tuple[str, str], dict[str, float]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    values: dict[tuple[str, str], dict[str, float]] = {}

    for ws in wb.worksheets:
        sheet_date = parse_date(ws.title)
        if sheet_date is None:
            match = re.fullmatch(r"May\s+(\d{2})", ws.title)
            if match:
                sheet_date = date(2026, 5, int(match.group(1)))
        if sheet_date is None or not (start <= sheet_date <= end):
            continue

        rows = list(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 31), values_only=True))
        raa_idx = None
        for idx, row in enumerate(rows):
            if row and norm_text(row[0]).lower() == "register audit adjustment":
                raa_idx = idx
                break
        if raa_idx is None:
            continue

        store_header = rows[0]
        raa_row = rows[raa_idx]
        for col_idx, value in enumerate(store_header):
            if not (isinstance(value, str) and value.strip().lower().startswith("store")):
                continue
            store = norm_store(value)
            if not store:
                continue
            values[(store, sheet_date.isoformat())] = {
                "ct_debit": number(raa_row[col_idx] if col_idx < len(raa_row) else None),
                "ct_credit": number(raa_row[col_idx + 1] if col_idx + 1 < len(raa_row) else None),
                "tqsr_debit": number(raa_row[col_idx + 2] if col_idx + 2 < len(raa_row) else None),
                "tqsr_credit": number(raa_row[col_idx + 3] if col_idx + 3 < len(raa_row) else None),
            }

    return values


def build_raa_check(
    raa_source: Path,
    comparison: Path,
    start: date,
    end: date,
    close_tolerance: float,
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    source_rows = load_raa_source_rows(raa_source, start, end)
    comparison_values = load_comparison_raa_values(comparison, start, end)

    rows: list[dict[str, Any]] = []
    for source in source_rows:
        found = comparison_values.get((source["store"], source["date"]))
        side = source["side"]
        if found:
            ct_amount = found["ct_debit"] if side == "Debit" else found["ct_credit"]
            tqsr_amount = found["tqsr_debit"] if side == "Debit" else found["tqsr_credit"]
            has_raa = abs(ct_amount) > 0.005
            variance = round(ct_amount - source["tqsr_source"], 2)
            abs_variance = round(abs(variance), 2)
            exact_match = abs_variance < 0.005
            close_match = abs_variance <= close_tolerance
            if has_raa and close_match:
                status = "In RAA - close"
            elif has_raa:
                status = "In RAA - not close"
            else:
                status = "No CenTech RAA value"
        else:
            found = {"ct_debit": 0.0, "ct_credit": 0.0, "tqsr_debit": 0.0, "tqsr_credit": 0.0}
            ct_amount = None
            tqsr_amount = None
            variance = None
            abs_variance = None
            has_raa = False
            exact_match = False
            close_match = False
            status = "Store/date not found in sales workbook"

        rows.append(
            {
                **source,
                "ct_debit": found["ct_debit"],
                "ct_credit": found["ct_credit"],
                "sales_tqsr_debit": found["tqsr_debit"],
                "sales_tqsr_credit": found["tqsr_credit"],
                "sales_tqsr_side": tqsr_amount,
                "ct_side": ct_amount,
                "variance_ct_vs_source_tqsr": variance,
                "abs_variance": abs_variance,
                "has_raa": has_raa,
                "exact_match": exact_match,
                "close_match": close_match,
                "status": status,
            }
        )

    summary = {
        "source_rows": len(rows),
        "unique_source_store_dates": len({(row["store"], row["date"]) for row in rows}),
        "sales_raa_store_dates_available": len(comparison_values),
        "rows_with_sales_store_date": sum(
            1 for row in rows if row["status"] != "Store/date not found in sales workbook"
        ),
        "rows_in_raa_now": sum(1 for row in rows if row["has_raa"]),
        "exact_matches": sum(1 for row in rows if row["exact_match"]),
        "close_matches": sum(1 for row in rows if row["close_match"]),
        "not_close_but_in_raa": sum(
            1 for row in rows if row["has_raa"] and not row["close_match"]
        ),
        "missing_ct_raa_value": sum(1 for row in rows if row["status"] == "No CenTech RAA value"),
        "missing_store_date": sum(
            1 for row in rows if row["status"] == "Store/date not found in sales workbook"
        ),
        "source_tqsr_total": round(sum(row["tqsr_source"] for row in rows), 2),
        "ct_raa_total_for_source_rows": round(sum((row["ct_side"] or 0) for row in rows), 2),
        "source_tqsr_total_for_rows_in_raa_now": round(
            sum(row["tqsr_source"] for row in rows if row["has_raa"]), 2
        ),
        "source_tqsr_total_for_missing_rows": round(
            sum(row["tqsr_source"] for row in rows if not row["has_raa"]), 2
        ),
        "status_counts": dict(Counter(row["status"] for row in rows)),
    }
    return summary, rows


def load_cash_account_mapping(path: Path) -> tuple[dict[tuple[str, str], tuple[str, str]], set[str], int]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = header_indexes(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    required = {"Class", "Transaction Category", "Account Number", "Account Name"}
    missing = sorted(required - set(headers))
    if missing:
        raise ValueError(f"{path}: missing headers: {missing}")

    mapping: dict[tuple[str, str], tuple[str, str]] = {}
    duplicate_keys = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        store = norm_store(row[headers["Class"]])
        category = norm_text(row[headers["Transaction Category"]])
        account_number = norm_account(row[headers["Account Number"]])
        account_name = norm_text(row[headers["Account Name"]])
        if not store or not category:
            continue
        key = (store, category)
        if key in mapping and mapping[key] != (account_number, account_name):
            duplicate_keys += 1
        mapping[key] = (account_number, account_name)

    categories = {category for _, category in mapping}
    return mapping, categories, duplicate_keys


def normalize_memo_category(raw: Any) -> str:
    category = norm_text(raw)
    category = re.sub(r"^CenTech\s*-\s*", "", category, flags=re.IGNORECASE)
    rewrites = {
        "Online Credit card": "Online Credit Card",
        "Pay-in": "Payin",
        "3rd Party - Grubhub": "3rd Party - GrubHub",
    }
    return rewrites.get(category, category)


def find_latest_sage_export() -> Path | None:
    candidates = [
        path
        for path in REPO_ROOT.glob("gl_journal_entries*")
        if path.is_file() and path.suffix.lower() in {".csv", ".xlsx", ".xlsm"}
    ]
    if not candidates:
        return None
    return max(candidates, key=lambda path: path.stat().st_mtime)


def iter_sage_rows(path: Path) -> tuple[list[str], Any]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        handle = path.open("r", encoding="utf-8-sig", newline="")
        reader = csv.DictReader(handle)
        headers = [norm_text(name).upper() for name in (reader.fieldnames or [])]

        def generator() -> Any:
            try:
                for row in reader:
                    yield {norm_text(key).upper(): value for key, value in row.items()}
            finally:
                handle.close()

        return headers, generator()

    if suffix in {".xlsx", ".xlsm"}:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [norm_text(value).upper() for value in header_row]

        def generator() -> Any:
            for values in ws.iter_rows(min_row=2, values_only=True):
                yield {
                    header: values[idx] if idx < len(values) else None
                    for idx, header in enumerate(headers)
                    if header
                }

        return headers, generator()

    raise ValueError(f"Unsupported Sage export format: {path}")


def build_sage_account_check(
    cash_accounts: Path,
    sage_export: Path,
    start: date,
    end: date,
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    mapping, mapped_categories, duplicate_keys = load_cash_account_mapping(cash_accounts)
    headers, sage_rows = iter_sage_rows(sage_export)
    required = {"DATE", "LOCATION_ID", "MEMO", "ACCT_NO", "DEBIT", "CREDIT"}
    missing = sorted(required - set(headers))
    if missing:
        raise ValueError(f"{sage_export}: missing required Sage columns: {missing}")

    grouped: dict[tuple[str, str], dict[str, Any]] = defaultdict(
        lambda: {
            "row_count": 0,
            "accounts": defaultdict(lambda: {"row_count": 0, "debit_total": 0.0, "credit_total": 0.0}),
        }
    )
    unmapped: dict[tuple[str, str], dict[str, float | int]] = defaultdict(
        lambda: {"row_count": 0, "debit_total": 0.0, "credit_total": 0.0}
    )
    rows_scanned = 0
    rows_in_mapped_categories = 0

    for row in sage_rows:
        row_date = parse_date(row.get("DATE"))
        if row_date is None or not (start <= row_date <= end):
            continue
        rows_scanned += 1

        category = normalize_memo_category(row.get("MEMO"))
        if category not in mapped_categories:
            continue
        rows_in_mapped_categories += 1

        store = norm_store(row.get("LOCATION_ID")) or ""
        account_number = norm_account(row.get("ACCT_NO"))
        debit = number(row.get("DEBIT"))
        credit = number(row.get("CREDIT"))
        key = (store, category)

        if key not in mapping:
            bucket = unmapped[key]
            bucket["row_count"] += 1
            bucket["debit_total"] += debit
            bucket["credit_total"] += credit
            continue

        bucket = grouped[key]
        bucket["row_count"] += 1
        account_bucket = bucket["accounts"][account_number]
        account_bucket["row_count"] += 1
        account_bucket["debit_total"] += debit
        account_bucket["credit_total"] += credit

    details: list[dict[str, Any]] = []
    matched_rows = 0
    mismatched_rows = 0
    missing_actual_mapped = 0

    def sort_key(key: tuple[str, str]) -> tuple[int | str, str]:
        store, category = key
        return (int(store) if store.isdigit() else store, category)

    for key in sorted(mapping, key=sort_key):
        store, category = key
        expected_number, _expected_name = mapping[key]
        bucket = grouped.get(key)
        if not bucket or not bucket["row_count"]:
            details.append(
                {
                    "store": store,
                    "category": category,
                    "expected_account_number": expected_number,
                    "actual_account_number": "",
                    "row_count": 0,
                    "debit_total": 0.0,
                    "credit_total": 0.0,
                    "number_match": "",
                    "status": "No Sage rows for mapped store/category",
                }
            )
            missing_actual_mapped += 1
            continue

        for actual_number, account_bucket in sorted(bucket["accounts"].items()):
            number_match = actual_number == expected_number
            status = "Match" if number_match else "Mismatch"
            if number_match:
                matched_rows += account_bucket["row_count"]
            else:
                mismatched_rows += account_bucket["row_count"]
            details.append(
                {
                    "store": store,
                    "category": category,
                    "expected_account_number": expected_number,
                    "actual_account_number": actual_number,
                    "row_count": account_bucket["row_count"],
                    "debit_total": round(account_bucket["debit_total"], 2),
                    "credit_total": round(account_bucket["credit_total"], 2),
                    "number_match": "Yes" if number_match else "No",
                    "status": status,
                }
            )

    for key in sorted(unmapped, key=sort_key):
        store, category = key
        bucket = unmapped[key]
        details.append(
            {
                "store": store,
                "category": category,
                "expected_account_number": "",
                "actual_account_number": "(multiple or unmapped)",
                "row_count": bucket["row_count"],
                "debit_total": round(float(bucket["debit_total"]), 2),
                "credit_total": round(float(bucket["credit_total"]), 2),
                "number_match": "No",
                "status": "Sage category/store missing from TQSR mapping",
            }
        )
        mismatched_rows += int(bucket["row_count"])

    summary = {
        "sage_export": str(sage_export),
        "expected_mappings": len(mapping),
        "mapped_categories": ", ".join(sorted(mapped_categories)),
        "sage_rows_scanned": rows_scanned,
        "sage_rows_in_mapped_categories": rows_in_mapped_categories,
        "detail_groups": len(details),
        "groups_matching": sum(1 for row in details if row["status"] == "Match"),
        "groups_mismatch": sum(1 for row in details if row["status"] == "Mismatch"),
        "mapped_groups_with_no_sage_rows": missing_actual_mapped,
        "sage_rows_matching_expected": matched_rows,
        "sage_rows_mismatching_expected": mismatched_rows,
        "duplicate_expected_mapping_keys": duplicate_keys,
        "status_counts": dict(Counter(row["status"] for row in details)),
    }
    return summary, details


def build_cash_account_check(
    cash_accounts: Path,
    centech_export: Path,
    start: date,
    end: date,
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    mapping, mapped_categories, duplicate_keys = load_cash_account_mapping(cash_accounts)

    wb = load_workbook(centech_export, read_only=True, data_only=True)
    ws = wb.active
    headers = header_indexes(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    required = {
        "Date",
        "Class",
        "Transaction Category",
        "Account Number",
        "Account Name",
        "Debit",
        "Credit",
    }
    missing = sorted(required - set(headers))
    if missing:
        raise ValueError(f"{centech_export}: missing headers: {missing}")

    grouped: dict[tuple[str, str], dict[str, Any]] = defaultdict(
        lambda: {
            "row_count": 0,
            "accounts": defaultdict(lambda: {"row_count": 0, "debit_total": 0.0, "credit_total": 0.0}),
        }
    )
    unmapped: dict[tuple[str, str], dict[str, float | int]] = defaultdict(
        lambda: {"row_count": 0, "debit_total": 0.0, "credit_total": 0.0}
    )
    rows_scanned = 0
    rows_in_mapped_categories = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_date = parse_date(row[headers["Date"]])
        if row_date is None or not (start <= row_date <= end):
            continue
        rows_scanned += 1

        category = norm_text(row[headers["Transaction Category"]])
        if category not in mapped_categories:
            continue
        rows_in_mapped_categories += 1

        store = norm_store(row[headers["Class"]])
        account_number = norm_account(row[headers["Account Number"]])
        account_name = norm_text(row[headers["Account Name"]])
        debit = number(row[headers["Debit"]])
        credit = number(row[headers["Credit"]])
        key = (store or "", category)

        if key not in mapping:
            bucket = unmapped[key]
            bucket["row_count"] += 1
            bucket["debit_total"] += debit
            bucket["credit_total"] += credit
            continue

        bucket = grouped[key]
        bucket["row_count"] += 1
        account_bucket = bucket["accounts"][(account_number, account_name)]
        account_bucket["row_count"] += 1
        account_bucket["debit_total"] += debit
        account_bucket["credit_total"] += credit

    details: list[dict[str, Any]] = []
    matched_rows = 0
    mismatched_rows = 0
    missing_actual_mapped = 0

    def sort_key(key: tuple[str, str]) -> tuple[int | str, str]:
        store, category = key
        return (int(store) if store.isdigit() else store, category)

    for key in sorted(mapping, key=sort_key):
        store, category = key
        expected_number, expected_name = mapping[key]
        bucket = grouped.get(key)
        if not bucket or not bucket["row_count"]:
            details.append(
                {
                    "store": store,
                    "category": category,
                    "expected_account_number": expected_number,
                    "expected_account_name": expected_name,
                    "actual_account_number": "",
                    "actual_account_name": "",
                    "row_count": 0,
                    "debit_total": 0.0,
                    "credit_total": 0.0,
                    "number_match": "",
                    "name_match": "",
                    "status": "No CenTech rows for mapped store/category",
                }
            )
            missing_actual_mapped += 1
            continue

        for (actual_number, actual_name), account_bucket in sorted(bucket["accounts"].items()):
            number_match = actual_number == expected_number
            name_match = actual_name == expected_name
            status = "Match" if number_match and name_match else "Mismatch"
            if status == "Match":
                matched_rows += account_bucket["row_count"]
            else:
                mismatched_rows += account_bucket["row_count"]
            details.append(
                {
                    "store": store,
                    "category": category,
                    "expected_account_number": expected_number,
                    "expected_account_name": expected_name,
                    "actual_account_number": actual_number,
                    "actual_account_name": actual_name,
                    "row_count": account_bucket["row_count"],
                    "debit_total": round(account_bucket["debit_total"], 2),
                    "credit_total": round(account_bucket["credit_total"], 2),
                    "number_match": "Yes" if number_match else "No",
                    "name_match": "Yes" if name_match else "No",
                    "status": status,
                }
            )

    for key in sorted(unmapped, key=sort_key):
        store, category = key
        bucket = unmapped[key]
        details.append(
            {
                "store": store,
                "category": category,
                "expected_account_number": "",
                "expected_account_name": "",
                "actual_account_number": "(multiple or unmapped)",
                "actual_account_name": "",
                "row_count": bucket["row_count"],
                "debit_total": round(float(bucket["debit_total"]), 2),
                "credit_total": round(float(bucket["credit_total"]), 2),
                "number_match": "No",
                "name_match": "No",
                "status": "CenTech category/store missing from TQSR mapping",
            }
        )
        mismatched_rows += int(bucket["row_count"])

    summary = {
        "expected_mappings": len(mapping),
        "mapped_categories": ", ".join(sorted(mapped_categories)),
        "centech_rows_scanned": rows_scanned,
        "centech_rows_in_mapped_categories": rows_in_mapped_categories,
        "detail_groups": len(details),
        "groups_matching": sum(1 for row in details if row["status"] == "Match"),
        "groups_mismatch": sum(1 for row in details if row["status"] == "Mismatch"),
        "mapped_groups_with_no_centech_rows": missing_actual_mapped,
        "centech_rows_matching_expected": matched_rows,
        "centech_rows_mismatching_expected": mismatched_rows,
        "duplicate_expected_mapping_keys": duplicate_keys,
        "status_counts": dict(Counter(row["status"] for row in details)),
    }
    return summary, details


class SheetStyles:
    title_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    metric_fill = PatternFill("solid", fgColor="EAF4EA")
    warn_fill = PatternFill("solid", fgColor="FCE4D6")
    ok_fill = PatternFill("solid", fgColor="E2F0D9")
    white_title = Font(color="FFFFFF", bold=True, size=14)
    sub_font = Font(color="666666", italic=True)
    bold = Font(bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    money_format = "$#,##0.00;[Red]($#,##0.00);-"


def write_metric_block(
    ws: Any,
    metrics: list[tuple[str, Any]],
    *,
    start_row: int = 4,
    styles: type[SheetStyles] = SheetStyles,
) -> None:
    for row_idx, (label, value) in enumerate(metrics, start_row):
        ws.cell(row_idx, 1, label)
        ws.cell(row_idx, 2, value)
        ws.cell(row_idx, 1).font = styles.bold
        ws.cell(row_idx, 1).fill = styles.metric_fill
        ws.cell(row_idx, 2).fill = styles.metric_fill
        ws.cell(row_idx, 1).border = styles.border
        ws.cell(row_idx, 2).border = styles.border
        if "total" in label.lower():
            ws.cell(row_idx, 2).number_format = styles.money_format


def write_status_counts(
    ws: Any,
    status_counts: dict[str, int],
    *,
    good_status: str,
    styles: type[SheetStyles] = SheetStyles,
) -> None:
    ws["D4"] = "Status"
    ws["E4"] = "Count"
    for cell in ws["D4:E4"][0]:
        cell.font = styles.bold
        cell.fill = styles.header_fill
        cell.border = styles.border
    for row_idx, (status, count) in enumerate(status_counts.items(), 5):
        ws.cell(row_idx, 4, status)
        ws.cell(row_idx, 5, count)
        ws.cell(row_idx, 4).border = styles.border
        ws.cell(row_idx, 5).border = styles.border
        fill = styles.ok_fill if status == good_status else styles.warn_fill
        ws.cell(row_idx, 4).fill = fill
        ws.cell(row_idx, 5).fill = fill


def create_raa_sheet(wb: Any, summary: dict[str, Any], rows: list[dict[str, Any]], close_tolerance: float) -> None:
    ws = wb.create_sheet("RAA Presence Check", 2 if "Discrepancies" in wb.sheetnames else 0)
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:N1")
    ws["A1"] = "Register Audit Adjustment Presence Check"
    ws["A1"].fill = SheetStyles.title_fill
    ws["A1"].font = SheetStyles.white_title
    ws.merge_cells("A2:N2")
    ws["A2"] = (
        "Compares Store + Date + Side from Centech_Register Audit vs Register Audit "
        f"Adjustment.xlsx to CenTech Register Audit Adjustment values. Close = absolute variance <= ${close_tolerance:,.2f}."
    )
    ws["A2"].font = SheetStyles.sub_font

    write_metric_block(
        ws,
        [
            ("Source rows checked", summary["source_rows"]),
            ("Unique source store/dates", summary["unique_source_store_dates"]),
            ("Rows with matching store/date in sales workbook", summary["rows_with_sales_store_date"]),
            ("Rows in RAA now", summary["rows_in_raa_now"]),
            (f"Close matches <= ${close_tolerance:,.2f}", summary["close_matches"]),
            ("Exact matches", summary["exact_matches"]),
            ("In RAA but not close", summary["not_close_but_in_raa"]),
            ("Missing CenTech RAA value", summary["missing_ct_raa_value"]),
            ("Store/date not found", summary["missing_store_date"]),
            ("Source TQSR total", summary["source_tqsr_total"]),
            ("CenTech RAA total for matched source rows", summary["ct_raa_total_for_source_rows"]),
            ("Source TQSR total for rows in RAA now", summary["source_tqsr_total_for_rows_in_raa_now"]),
            ("Source TQSR total for missing rows", summary["source_tqsr_total_for_missing_rows"]),
        ],
    )
    write_status_counts(ws, summary["status_counts"], good_status="In RAA - close")

    start_row = 19
    headers = [
        "Store",
        "Date",
        "Side",
        "Source TQSR Amount",
        "Sales RAA CenTech Debit",
        "Sales RAA CenTech Credit",
        "Sales RAA CenTech Side Amount",
        "Sales RAA TQSR Side Amount",
        "Variance: CenTech - Source TQSR",
        "Abs Variance",
        "Has RAA Now?",
        f"Close <= ${close_tolerance:,.2f}?",
        "Exact Match?",
        "Status",
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(start_row, col_idx, header)
        cell.font = SheetStyles.bold
        cell.fill = SheetStyles.header_fill
        cell.border = SheetStyles.border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row_idx, row in enumerate(rows, start_row + 1):
        values = [
            row["store"],
            row["date"],
            row["side"],
            row["tqsr_source"],
            row["ct_debit"],
            row["ct_credit"],
            row["ct_side"],
            row["sales_tqsr_side"],
            row["variance_ct_vs_source_tqsr"],
            row["abs_variance"],
            "Yes" if row["has_raa"] else "No",
            "Yes" if row["close_match"] else "No",
            "Yes" if row["exact_match"] else "No",
            row["status"],
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row_idx, col_idx, value)
            cell.border = SheetStyles.border
            if col_idx in (4, 5, 6, 7, 8, 9, 10):
                cell.number_format = SheetStyles.money_format
            if col_idx in (11, 12, 13):
                cell.alignment = Alignment(horizontal="center")
        if row["status"] != "In RAA - close":
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row_idx, col_idx).fill = SheetStyles.warn_fill

    end_row = start_row + len(rows)
    table = Table(displayName="RAAPresenceTable", ref=f"A{start_row}:N{end_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)
    ws.freeze_panes = f"A{start_row + 1}"
    widths = {
        "A": 10,
        "B": 12,
        "C": 10,
        "D": 18,
        "E": 20,
        "F": 20,
        "G": 22,
        "H": 22,
        "I": 24,
        "J": 14,
        "K": 13,
        "L": 16,
        "M": 13,
        "N": 34,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 32
    ws.row_dimensions[start_row].height = 34


def create_cash_account_sheet(wb: Any, summary: dict[str, Any], rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Cash Account Check", 3 if "RAA Presence Check" in wb.sheetnames else 0)
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:L1")
    ws["A1"] = "TQSR Cash Account Mapping Check"
    ws["A1"].fill = SheetStyles.title_fill
    ws["A1"].font = SheetStyles.white_title
    ws.merge_cells("A2:L2")
    ws["A2"] = (
        "Compares expected Class + Transaction Category account number/name from "
        "TQSR Cash Accounts.xlsx to CenTech export rows. Detail is grouped by store/category/account."
    )
    ws["A2"].font = SheetStyles.sub_font

    write_metric_block(
        ws,
        [
            ("Expected mapping rows", summary["expected_mappings"]),
            ("Mapped categories", summary["mapped_categories"]),
            ("CenTech rows scanned", summary["centech_rows_scanned"]),
            ("CenTech rows in mapped categories", summary["centech_rows_in_mapped_categories"]),
            ("Detail groups", summary["detail_groups"]),
            ("Groups matching", summary["groups_matching"]),
            ("Groups with account mismatch", summary["groups_mismatch"]),
            ("Mapped groups with no CenTech rows", summary["mapped_groups_with_no_centech_rows"]),
            ("CenTech rows matching expected", summary["centech_rows_matching_expected"]),
            ("CenTech rows mismatching expected", summary["centech_rows_mismatching_expected"]),
            ("Duplicate expected mapping keys", summary["duplicate_expected_mapping_keys"]),
        ],
    )
    write_status_counts(ws, summary["status_counts"], good_status="Match")

    start_row = 18
    headers = [
        "Store",
        "Transaction Category",
        "Expected Account Number",
        "Expected Account Name",
        "CenTech Account Number",
        "CenTech Account Name",
        "CenTech Row Count",
        "Debit Total",
        "Credit Total",
        "Account Number Match?",
        "Account Name Match?",
        "Status",
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(start_row, col_idx, header)
        cell.font = SheetStyles.bold
        cell.fill = SheetStyles.header_fill
        cell.border = SheetStyles.border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row_idx, row in enumerate(rows, start_row + 1):
        values = [
            row["store"],
            row["category"],
            row["expected_account_number"],
            row["expected_account_name"],
            row["actual_account_number"],
            row["actual_account_name"],
            row["row_count"],
            row["debit_total"],
            row["credit_total"],
            row["number_match"],
            row["name_match"],
            row["status"],
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row_idx, col_idx, value)
            cell.border = SheetStyles.border
            if col_idx in (8, 9):
                cell.number_format = SheetStyles.money_format
            if col_idx in (7, 10, 11):
                cell.alignment = Alignment(horizontal="center")
        if row["status"] != "Match":
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row_idx, col_idx).fill = SheetStyles.warn_fill

    end_row = start_row + len(rows)
    table = Table(displayName="CashAccountCheckTable", ref=f"A{start_row}:L{end_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)
    ws.freeze_panes = f"A{start_row + 1}"
    widths = {
        "A": 10,
        "B": 24,
        "C": 20,
        "D": 38,
        "E": 20,
        "F": 42,
        "G": 14,
        "H": 14,
        "I": 14,
        "J": 18,
        "K": 18,
        "L": 42,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 32
    ws.row_dimensions[start_row].height = 36


def create_sage_account_sheet(wb: Any, summary: dict[str, Any], rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Sage Account Check", 4 if "Cash Account Check" in wb.sheetnames else 0)
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:I1")
    ws["A1"] = "Sage GL Account Mapping Check"
    ws["A1"].fill = SheetStyles.title_fill
    ws["A1"].font = SheetStyles.white_title
    ws.merge_cells("A2:I2")
    ws["A2"] = (
        "Compares expected Class + Transaction Category account number from TQSR Cash Accounts.xlsx "
        "to Sage GL journal rows. Sage category is parsed from MEMO after removing the Centech - prefix."
    )
    ws["A2"].font = SheetStyles.sub_font

    write_metric_block(
        ws,
        [
            ("Sage export file", Path(summary["sage_export"]).name),
            ("Expected mapping rows", summary["expected_mappings"]),
            ("Mapped categories", summary["mapped_categories"]),
            ("Sage rows scanned", summary["sage_rows_scanned"]),
            ("Sage rows in mapped categories", summary["sage_rows_in_mapped_categories"]),
            ("Detail groups", summary["detail_groups"]),
            ("Groups matching", summary["groups_matching"]),
            ("Groups with account mismatch", summary["groups_mismatch"]),
            ("Mapped groups with no Sage rows", summary["mapped_groups_with_no_sage_rows"]),
            ("Sage rows matching expected", summary["sage_rows_matching_expected"]),
            ("Sage rows mismatching expected", summary["sage_rows_mismatching_expected"]),
            ("Duplicate expected mapping keys", summary["duplicate_expected_mapping_keys"]),
        ],
    )
    write_status_counts(ws, summary["status_counts"], good_status="Match")

    start_row = 19
    headers = [
        "Store",
        "Transaction Category",
        "Expected Account Number",
        "Sage ACCT_NO",
        "Sage Row Count",
        "Debit Total",
        "Credit Total",
        "Account Number Match?",
        "Status",
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(start_row, col_idx, header)
        cell.font = SheetStyles.bold
        cell.fill = SheetStyles.header_fill
        cell.border = SheetStyles.border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row_idx, row in enumerate(rows, start_row + 1):
        values = [
            row["store"],
            row["category"],
            row["expected_account_number"],
            row["actual_account_number"],
            row["row_count"],
            row["debit_total"],
            row["credit_total"],
            row["number_match"],
            row["status"],
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row_idx, col_idx, value)
            cell.border = SheetStyles.border
            if col_idx in (6, 7):
                cell.number_format = SheetStyles.money_format
            if col_idx in (5, 8):
                cell.alignment = Alignment(horizontal="center")
        if row["status"] != "Match":
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row_idx, col_idx).fill = SheetStyles.warn_fill

    end_row = start_row + len(rows)
    table = Table(displayName="SageAccountCheckTable", ref=f"A{start_row}:I{end_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)
    ws.freeze_panes = f"A{start_row + 1}"
    widths = {
        "A": 10,
        "B": 24,
        "C": 22,
        "D": 18,
        "E": 14,
        "F": 14,
        "G": 14,
        "H": 20,
        "I": 42,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 32
    ws.row_dimensions[start_row].height = 36


def default_output_path(comparison: Path) -> Path:
    if comparison.stem.endswith("_with_RAA_check"):
        return comparison
    return comparison.with_name(f"{comparison.stem}_with_RAA_check{comparison.suffix}")


def write_workbook(
    comparison: Path,
    output: Path,
    raa_summary: dict[str, Any],
    raa_rows: list[dict[str, Any]],
    cash_summary: dict[str, Any],
    cash_rows: list[dict[str, Any]],
    sage_summary: dict[str, Any] | None,
    sage_rows: list[dict[str, Any]] | None,
    close_tolerance: float,
) -> None:
    if comparison.resolve() != output.resolve():
        shutil.copy2(comparison, output)
    wb = load_workbook(output)
    for sheet_name in ("RAA Presence Check", "Cash Account Check", "Sage Account Check"):
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
    create_raa_sheet(wb, raa_summary, raa_rows, close_tolerance)
    create_cash_account_sheet(wb, cash_summary, cash_rows)
    if sage_summary is not None and sage_rows is not None:
        create_sage_account_sheet(wb, sage_summary, sage_rows)
    wb.save(output)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Add RAA presence and TQSR cash account check tabs to a sales comparison workbook."
    )
    parser.add_argument("--start", type=parse_cli_date, required=True)
    parser.add_argument("--end", type=parse_cli_date, required=True)
    parser.add_argument(
        "--comparison",
        type=Path,
        default=DEFAULT_RUN_ROOT
        / "output"
        / "Sales_CenTech_vs_Client_2026-05-01_2026-05-31.xlsx",
    )
    parser.add_argument(
        "--centech-export",
        type=Path,
        default=DEFAULT_RUN_ROOT / "input" / "centech_export.xlsx",
    )
    parser.add_argument(
        "--raa-source",
        type=Path,
        default=REPO_ROOT / "Centech_Register Audit vs Register Audit Adjustment.xlsx",
    )
    parser.add_argument(
        "--cash-accounts",
        type=Path,
        default=REPO_ROOT / "TQSR Cash Accounts.xlsx",
    )
    parser.add_argument(
        "--sage-export",
        type=Path,
        help="Sage GL journal export. Defaults to newest repo-root file matching gl_journal_entries*.",
    )
    parser.add_argument(
        "--skip-sage",
        action="store_true",
        help="Do not add the Sage Account Check tab.",
    )
    parser.add_argument("--output", type=Path)
    parser.add_argument("--close-tolerance", type=float, default=1.0)
    parser.add_argument(
        "--summary-json",
        type=Path,
        help="Optional path for a compact JSON summary of both checks.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if args.start > args.end:
        raise SystemExit("--start must be on or before --end")

    comparison = args.comparison.resolve()
    centech_export = args.centech_export.resolve()
    raa_source = args.raa_source.resolve()
    cash_accounts = args.cash_accounts.resolve()
    sage_export = args.sage_export.resolve() if args.sage_export else find_latest_sage_export()
    output = (args.output or default_output_path(comparison)).resolve()

    for path in (comparison, centech_export, raa_source, cash_accounts):
        if not path.exists():
            raise SystemExit(f"Missing required file: {path}")
    if not args.skip_sage and sage_export is None:
        raise SystemExit(
            "No Sage export found. Put gl_journal_entries* in the repo root, "
            "pass --sage-export, or use --skip-sage."
        )
    if not args.skip_sage and sage_export is not None and not sage_export.exists():
        raise SystemExit(f"Missing Sage export file: {sage_export}")

    raa_summary, raa_rows = build_raa_check(
        raa_source, comparison, args.start, args.end, args.close_tolerance
    )
    cash_summary, cash_rows = build_cash_account_check(
        cash_accounts, centech_export, args.start, args.end
    )
    sage_summary = None
    sage_rows = None
    if not args.skip_sage and sage_export is not None:
        sage_summary, sage_rows = build_sage_account_check(
            cash_accounts, sage_export, args.start, args.end
        )
    write_workbook(
        comparison,
        output,
        raa_summary,
        raa_rows,
        cash_summary,
        cash_rows,
        sage_summary,
        sage_rows,
        args.close_tolerance,
    )

    result = {
        "output": str(output),
        "raa_summary": raa_summary,
        "cash_account_summary": cash_summary,
        "sage_account_summary": sage_summary,
    }
    if args.summary_json:
        args.summary_json.parent.mkdir(parents=True, exist_ok=True)
        args.summary_json.write_text(json.dumps(result, indent=2), encoding="utf-8")
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
