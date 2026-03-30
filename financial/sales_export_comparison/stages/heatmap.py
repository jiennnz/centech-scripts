from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, PatternFill

from financial.sales_export_comparison.stages.template_builder import TableLayout


CT_OR_CLIENT_ONLY_FILL = PatternFill(fill_type="solid", start_color="FFEB9C", end_color="FFEB9C")
MISSING_FILL = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")
WRONG_SIDE_FILL = PatternFill(fill_type="solid", start_color="FFB6C1", end_color="FFB6C1")


@dataclass(frozen=True)
class HeatmapConfig:
    workbook_path: Path
    stores: list[str]
    category_rows: dict[str, int]
    tolerance: float = 0.0
    layout: TableLayout = TableLayout()


def _col_letter(n: int) -> str:
    result = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


def _to_number(value: object) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "").strip())
    except ValueError:
        return 0.0


def _build_store_mapping(ws, stores: list[str], layout: TableLayout) -> dict[str, dict[str, int]]:
    mapping: dict[str, dict[str, int]] = {}
    for idx, store in enumerate(stores):
        start_col = layout.start_col + (idx * layout.block_width)
        mapping[str(store)] = {
            "ct_debit": start_col,
            "ct_credit": start_col + 1,
            "client_debit": start_col + 2,
            "client_credit": start_col + 3,
        }
    return mapping


def _build_final_discrepancies(raw: dict[str, dict[str, dict[str, list]]]) -> dict[str, list[dict]]:
    final: dict[str, list[dict]] = {date_sheet: [] for date_sheet in raw}
    for date_sheet, store_buckets in raw.items():
        for store_num, bucket in store_buckets.items():
            if bucket["ct_only"]:
                cats = bucket["ct_only"]
                final[date_sheet].append(
                    {
                        "store": store_num,
                        "category": f"{len(cats)} categor{'y' if len(cats) == 1 else 'ies'}: {', '.join(cats)}",
                        "ct_debit": "-",
                        "ct_credit": "-",
                        "client_debit": "-",
                        "client_credit": "-",
                        "variance": "-",
                        "issue_type": "CT Only - No Client Data",
                        "color": "ct_or_client_only",
                    }
                )
            if bucket["client_only"]:
                cats = bucket["client_only"]
                final[date_sheet].append(
                    {
                        "store": store_num,
                        "category": f"{len(cats)} categor{'y' if len(cats) == 1 else 'ies'}: {', '.join(cats)}",
                        "ct_debit": "-",
                        "ct_credit": "-",
                        "client_debit": "-",
                        "client_credit": "-",
                        "variance": "-",
                        "issue_type": "Client Only - No CT Data",
                        "color": "ct_or_client_only",
                    }
                )
            final[date_sheet].extend(bucket["mismatches"])
    return final


def run(config: HeatmapConfig) -> Path:
    wb = load_workbook(config.workbook_path)
    wb_read = load_workbook(config.workbook_path, data_only=True)

    if "Heatmap" in wb.sheetnames:
        wb.remove(wb["Heatmap"])
    if "Discrepancies" in wb.sheetnames:
        wb.remove(wb["Discrepancies"])

    date_sheets = [s for s in wb.sheetnames if s not in {"Heatmap", "Discrepancies"}]
    categories = list(config.category_rows.keys())

    heatmap = wb.create_sheet("Heatmap", 0)
    heatmap["A1"] = "Category Match Heatmap"
    heatmap["A1"].font = Font(bold=True, size=14)
    heatmap["A3"] = "Green = 100% match, Yellow = 50% match, Red = 0% match"
    heatmap["C5"] = "Category"
    heatmap["C5"].font = Font(bold=True)

    for idx, sheet_name in enumerate(date_sheets):
        col = _col_letter(4 + idx)
        heatmap[f"{col}5"] = sheet_name
        heatmap[f"{col}5"].font = Font(bold=True)

    for idx, category in enumerate(categories):
        heatmap[f"C{6 + idx}"] = category

    raw: dict[str, dict[str, dict[str, list]]] = {date_sheet: {} for date_sheet in date_sheets}

    for date_idx, date_sheet in enumerate(date_sheets):
        ws = wb[date_sheet]
        ws_read = wb_read[date_sheet]
        store_mapping = _build_store_mapping(ws, config.stores, config.layout)
        data_col = _col_letter(4 + date_idx)

        for cat_idx, category in enumerate(categories):
            row_num = int(config.category_rows[category])
            out_row = 6 + cat_idx
            is_raa = category == "Register Audit Adjustment"

            matches = 0
            stores_with_client_data = 0

            for store in config.stores:
                store_key = str(store)
                cols = store_mapping.get(store_key)
                if not cols:
                    continue

                ct_debit = _to_number(ws_read.cell(row=row_num, column=cols["ct_debit"]).value)
                ct_credit = _to_number(ws_read.cell(row=row_num, column=cols["ct_credit"]).value)
                cl_debit = _to_number(ws_read.cell(row=row_num, column=cols["client_debit"]).value)
                cl_credit = _to_number(ws_read.cell(row=row_num, column=cols["client_credit"]).value)

                ct_has_debit = abs(ct_debit) > config.tolerance
                ct_has_credit = abs(ct_credit) > config.tolerance
                cl_has_debit = abs(cl_debit) > config.tolerance
                cl_has_credit = abs(cl_credit) > config.tolerance
                ct_has_data = ct_has_debit or ct_has_credit
                cl_has_data = cl_has_debit or cl_has_credit

                if store_key not in raw[date_sheet]:
                    raw[date_sheet][store_key] = {"ct_only": [], "client_only": [], "mismatches": []}
                bucket = raw[date_sheet][store_key]

                if not ct_has_data and not cl_has_data:
                    continue

                if ct_has_data and not cl_has_data and not is_raa:
                    bucket["ct_only"].append(category)
                    continue

                if cl_has_data and not ct_has_data and not is_raa:
                    bucket["client_only"].append(category)
                    stores_with_client_data += 1
                    continue

                stores_with_client_data += 1
                is_match = True
                issue_details: list[str] = []
                color: str | None = None

                if is_raa:
                    if ct_has_data and not cl_has_data:
                        is_match = False
                        color = "ct_or_client_only"
                        issue_details.append("RAA: CenTech Has Data (Client Empty)")
                    elif cl_has_data and not ct_has_data:
                        is_match = True
                    else:
                        if cl_has_debit:
                            if not ct_has_debit:
                                issue_details.append("RAA: Client debit but CT missing/on credit")
                                is_match = False
                            elif abs(ct_debit - cl_debit) > config.tolerance:
                                issue_details.append(f"RAA: Debit mismatch (${abs(ct_debit - cl_debit):.2f})")
                                is_match = False
                        if cl_has_credit:
                            if not ct_has_credit:
                                issue_details.append("RAA: Client credit but CT missing/on debit")
                                is_match = False
                            elif abs(ct_credit - cl_credit) > config.tolerance:
                                issue_details.append(f"RAA: Credit mismatch (${abs(ct_credit - cl_credit):.2f})")
                                is_match = False
                else:
                    if cl_has_debit:
                        if not ct_has_debit:
                            issue_details.append("CT missing debit")
                            is_match = False
                        elif abs(ct_debit - cl_debit) > config.tolerance:
                            issue_details.append(f"Debit: CT ${ct_debit:.2f} vs Client ${cl_debit:.2f}")
                            is_match = False
                        if ct_has_credit:
                            issue_details.append("CT has credit when Client has debit (wrong side)")
                            is_match = False

                    if cl_has_credit:
                        if not ct_has_credit:
                            issue_details.append("CT missing credit")
                            is_match = False
                        elif abs(ct_credit - cl_credit) > config.tolerance:
                            issue_details.append(f"Credit: CT ${ct_credit:.2f} vs Client ${cl_credit:.2f}")
                            is_match = False
                        if ct_has_debit:
                            issue_details.append("CT has debit when Client has credit (wrong side)")
                            is_match = False

                    if not is_match:
                        issue_text = " ".join(issue_details).lower()
                        if "missing" in issue_text:
                            color = "missing"
                        elif "wrong side" in issue_text:
                            color = "wrong_side"

                if is_match:
                    matches += 1
                else:
                    bucket["mismatches"].append(
                        {
                            "store": store_key,
                            "category": category,
                            "ct_debit": ct_debit,
                            "ct_credit": ct_credit,
                            "client_debit": cl_debit,
                            "client_credit": cl_credit,
                            "variance": (ct_debit + ct_credit) - (cl_debit + cl_credit),
                            "issue_type": "; ".join(issue_details),
                            "color": color,
                        }
                    )

            if stores_with_client_data > 0:
                heatmap[f"{data_col}{out_row}"] = matches / stores_with_client_data
                heatmap[f"{data_col}{out_row}"].number_format = "0%"
            else:
                heatmap[f"{data_col}{out_row}"] = "N/A"

    if date_sheets and categories:
        end_col = _col_letter(4 + len(date_sheets) - 1)
        end_row = 6 + len(categories) - 1
        data_range = f"D6:{end_col}{end_row}"
        rule = ColorScaleRule(
            start_type="num",
            start_value=0,
            start_color="F8696B",
            mid_type="num",
            mid_value=0.5,
            mid_color="FFEB84",
            end_type="num",
            end_value=1,
            end_color="63BE7B",
        )
        heatmap.conditional_formatting.add(data_range, rule)

    disc_sheet = wb.create_sheet("Discrepancies")
    disc_sheet["A1"] = "All Discrepancies"
    disc_sheet["A1"].font = Font(bold=True, size=16)
    current_row = 3

    date_discrepancies = _build_final_discrepancies(raw)
    for date_sheet in date_sheets:
        discrepancies = date_discrepancies[date_sheet]
        if not discrepancies:
            continue

        disc_sheet[f"A{current_row}"] = f"{date_sheet} - {len(discrepancies)} discrepancies"
        disc_sheet[f"A{current_row}"].font = Font(bold=True, size=12, color="FFFFFF")
        disc_sheet[f"A{current_row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        disc_sheet.merge_cells(f"A{current_row}:H{current_row}")
        current_row += 1

        headers = ["Store", "Category", "CT Debit", "CT Credit", "Client Debit", "Client Credit", "Net Variance", "Issue Type"]
        for col_idx, header in enumerate(headers, start=1):
            c = disc_sheet.cell(row=current_row, column=col_idx, value=header)
            c.font = Font(bold=True)
            c.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        current_row += 1

        for disc in discrepancies:
            disc_sheet.cell(row=current_row, column=1, value=disc["store"])
            disc_sheet.cell(row=current_row, column=2, value=disc["category"])
            for col_idx, key in [(3, "ct_debit"), (4, "ct_credit"), (5, "client_debit"), (6, "client_credit"), (7, "variance")]:
                cell = disc_sheet.cell(row=current_row, column=col_idx)
                val = disc[key]
                if val == "-":
                    cell.value = "-"
                else:
                    cell.value = val
                    cell.number_format = "$#,##0.00"
            disc_sheet.cell(row=current_row, column=8, value=disc["issue_type"])

            color = disc.get("color")
            fill = None
            if color == "ct_or_client_only":
                fill = CT_OR_CLIENT_ONLY_FILL
            elif color == "missing":
                fill = MISSING_FILL
            elif color == "wrong_side":
                fill = WRONG_SIDE_FILL
            if fill:
                for col_idx in range(1, 9):
                    disc_sheet.cell(row=current_row, column=col_idx).fill = fill
            current_row += 1

        current_row += 2

    disc_sheet.column_dimensions["A"].width = 10
    disc_sheet.column_dimensions["B"].width = 60
    disc_sheet.column_dimensions["C"].width = 12
    disc_sheet.column_dimensions["D"].width = 12
    disc_sheet.column_dimensions["E"].width = 12
    disc_sheet.column_dimensions["F"].width = 12
    disc_sheet.column_dimensions["G"].width = 12
    disc_sheet.column_dimensions["H"].width = 50

    heatmap.column_dimensions["A"].width = 5
    heatmap.column_dimensions["B"].width = 5
    heatmap.column_dimensions["C"].width = 30
    for idx in range(len(date_sheets)):
        heatmap.column_dimensions[_col_letter(4 + idx)].width = 12

    wb_read.close()
    wb.save(config.workbook_path)
    return config.workbook_path
