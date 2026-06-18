from __future__ import annotations

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from pathlib import Path

from payroll.stages.comparison.style import (
    autofit_columns,
    style_sheet,
    style_summary_row,
    apply_red_text,
    apply_alternating_color,
)


RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
HEADER_FILL = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
TITLE_FILL = PatternFill(start_color="000000", end_color="000000", fill_type="solid")


def create_workbook():
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    return wb


def create_discrepancy_collector() -> dict[str, list[dict]]:
    return {
        "missing_employee_numbers": [],
        "wrong_tips": [],
        "wrong_hours": [],
    }


def _to_float(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, str):
        value = value.replace("$", "").replace(",", "").strip()
        if not value:
            return 0.0
    try:
        if value != value:
            return 0.0
    except TypeError:
        pass
    return float(value)


def _employee_number(value) -> str:
    if value is None:
        return ""
    try:
        if value != value:
            return ""
    except TypeError:
        pass
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if text.endswith(".0"):
        return text[:-2]
    return text


def _is_missing_employee_number(value) -> bool:
    return _employee_number(value) in {"", "0"}


def _hours_key(regular_hours: float, overtime_hours: float) -> tuple[float, float]:
    return (round(regular_hours, 2), round(overtime_hours, 2))


def _has_missing_employee_pair(generated_record: dict, webapp_record: dict) -> bool:
    return (
        not _is_missing_employee_number(generated_record["employee_number"])
        and _is_missing_employee_number(webapp_record["employee_number"])
    )


def _has_same_hours(generated_record: dict, webapp_record: dict) -> bool:
    return _hours_key(
        generated_record["regular_hours"],
        generated_record["overtime_hours"],
    ) == _hours_key(
        webapp_record["regular_hours"],
        webapp_record["overtime_hours"],
    )


def _append_missing_employee_number(discrepancies: dict[str, list[dict]], store, generated_record: dict, webapp_record: dict) -> None:
    discrepancies["missing_employee_numbers"].append({
        "store": store,
        "centech_employee_number": _employee_number(webapp_record["employee_number"]) or "0",
        "qa_employee_number": _employee_number(generated_record["employee_number"]) or "0",
        "employee_name": generated_record.get("employee_name") or webapp_record.get("employee_name"),
        "centech_regular_hours": webapp_record["regular_hours"],
        "centech_overtime_hours": webapp_record["overtime_hours"],
        "qa_regular_hours": generated_record["regular_hours"],
        "qa_overtime_hours": generated_record["overtime_hours"],
    })


def _append_wrong_hours(discrepancies: dict[str, list[dict]], store, generated_record: dict, webapp_record: dict) -> None:
    employee_number = ""
    for record in [generated_record, webapp_record]:
        if not _is_missing_employee_number(record["employee_number"]):
            employee_number = _employee_number(record["employee_number"])
            break
    if not employee_number:
        employee_number = _employee_number(generated_record["employee_number"])
    discrepancies["wrong_hours"].append({
        "store": store,
        "employee_number": employee_number,
        "employee_name": generated_record.get("employee_name") or webapp_record.get("employee_name"),
        "qa_regular_hours": generated_record["regular_hours"],
        "qa_overtime_hours": generated_record["overtime_hours"],
        "centech_regular_hours": webapp_record["regular_hours"],
        "centech_overtime_hours": webapp_record["overtime_hours"],
    })


def _style_discrepancy_sheet(ws, title: str, headers: list[str], rows: list[list]) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 1))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.fill = TITLE_FILL

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_num, value=header)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL

    if rows:
        for row_num, row in enumerate(rows, start=4):
            for col_num, value in enumerate(row, start=1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = RED_FILL
    else:
        cell = ws.cell(row=4, column=1, value="No discrepancies found")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A4"
    autofit_columns(ws)


def add_discrepancy_sheets(wb, discrepancies: dict[str, list[dict]]) -> None:
    missing_rows = [
        [
            row["store"],
            row["centech_employee_number"],
            row["qa_employee_number"],
            row["employee_name"],
            row["centech_regular_hours"],
            row["centech_overtime_hours"],
            row["qa_regular_hours"],
            row["qa_overtime_hours"],
        ]
        for row in discrepancies["missing_employee_numbers"]
    ]
    wrong_tip_rows = [
        [row["store"], row["date"]]
        for row in discrepancies["wrong_tips"]
    ]
    wrong_hour_rows = [
        [
            row["store"],
            row["employee_number"],
            row["employee_name"],
            row["qa_regular_hours"],
            row["qa_overtime_hours"],
            row["centech_regular_hours"],
            row["centech_overtime_hours"],
        ]
        for row in discrepancies["wrong_hours"]
    ]

    sheet_specs = [
        (
            "Missing Emp Numbers",
            "Missing Employee Numbers",
            [
                "Store",
                "CenTech Employee Number",
                "QA Employee Number",
                "Employee Name",
                "CenTech Regular Hours",
                "CenTech Overtime Hours",
                "QA Regular Hours",
                "QA Overtime Hours",
            ],
            missing_rows,
            None,
        ),
        (
            "Wrong Tips",
            "Wrong Tips",
            ["Store", "Date"],
            wrong_tip_rows,
            "FFFF00",
        ),
        (
            "Wrong Hours",
            "Wrong Hours",
            [
                "Store",
                "Employee Number",
                "Employee Name",
                "QA Regular Hours",
                "QA Overtime Hours",
                "CenTech Regular Hours",
                "CenTech Overtime Hours",
            ],
            wrong_hour_rows,
            "FF0000",
        ),
    ]

    for index, (sheet_name, title, headers, rows, tab_color) in enumerate(sheet_specs):
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(title=sheet_name, index=index)
        if rows and tab_color:
            ws.sheet_properties.tabColor = tab_color
        _style_discrepancy_sheet(ws, title, headers, rows)


def add_store_data_to_sheet(
    wb,
    store,
    generated_store_data,
    webapp_store_data,
    centech_tips: dict | None = None,
    discrepancies: dict[str, list[dict]] | None = None,
    tip_date_label: str | None = None,
) -> None:
    generated_sorted = generated_store_data.sort_values(by="Employee Number")
    webapp_sorted = webapp_store_data.sort_values(by="Employee Number")

    ws = wb.create_sheet(title=str(store))
    style_sheet(ws, store_number=store)

    total_hours_generated = 0.0
    total_tip_generated = 0.0
    total_hours_webapp = 0.0
    total_tip_webapp = 0.0
    generated_data_map = {}
    generated_records = []
    webapp_records = []
    has_hours_discrepancy = False
    has_tips_discrepancy = False

    for index, row in enumerate(generated_sorted.iterrows(), start=5):
        employee_number = row[1]["Employee Number"]
        employee_name = row[1]["Employee Name"]
        regular_hours = _to_float(row[1]["Regular Hours"])
        overtime_hours = _to_float(row[1]["Overtime Hours"])
        coded_amount = row[1]["Coded Amount"]

        coded_amount = _to_float(coded_amount)

        total_hours = regular_hours + overtime_hours
        tip_rate = coded_amount / total_hours if total_hours > 0 else 0

        total_hours_generated += total_hours
        total_tip_generated += coded_amount

        ws.cell(row=index, column=1, value=employee_number)
        ws.cell(row=index, column=2, value=employee_name)
        ws.cell(row=index, column=3, value=regular_hours)
        ws.cell(row=index, column=4, value=overtime_hours)
        ws.cell(row=index, column=5, value=coded_amount)
        ws.cell(row=index, column=6, value=tip_rate)

        employee_key = _employee_number(employee_number)
        generated_record = {
            "row": index,
            "regular_hours": regular_hours,
            "overtime_hours": overtime_hours,
            "employee_name": employee_name,
            "employee_number": employee_number,
            "matched": False,
        }
        if not _is_missing_employee_number(employee_number):
            generated_data_map[employee_key] = generated_record
        generated_records.append(generated_record)

        for col_num in range(1, 7):
            ws.cell(row=index, column=col_num).alignment = Alignment(horizontal="center", vertical="center")

    for index, row in enumerate(webapp_sorted.iterrows(), start=5):
        employee_number = row[1]["Employee Number"]
        regular_hours = _to_float(row[1]["Regular Hours"])
        overtime_hours = _to_float(row[1]["Overtime Hours"])
        coded_amount = row[1]["Coded Amount"]

        coded_amount = _to_float(coded_amount)

        total_hours = regular_hours + overtime_hours
        tip_rate = coded_amount / total_hours if total_hours > 0 else 0

        total_hours_webapp += total_hours
        total_tip_webapp += coded_amount

        ws.cell(row=index, column=8, value=employee_number)
        ws.cell(row=index, column=9, value=regular_hours)
        ws.cell(row=index, column=10, value=overtime_hours)
        ws.cell(row=index, column=11, value=coded_amount)
        ws.cell(row=index, column=12, value=tip_rate)

        for col_num in range(8, 13):
            ws.cell(row=index, column=col_num).alignment = Alignment(horizontal="center", vertical="center")

        webapp_records.append({
            "row": index,
            "employee_number": employee_number,
            "regular_hours": regular_hours,
            "overtime_hours": overtime_hours,
            "matched": False,
        })

        employee_key = _employee_number(employee_number)
        if not _is_missing_employee_number(employee_number) and employee_key in generated_data_map:
            gen_row = generated_data_map[employee_key]["row"]
            gen_regular = generated_data_map[employee_key]["regular_hours"]
            gen_overtime = generated_data_map[employee_key]["overtime_hours"]
            generated_data_map[employee_key]["matched"] = True
            webapp_records[-1]["matched"] = True

            if regular_hours != gen_regular or overtime_hours != gen_overtime:
                has_hours_discrepancy = True
                apply_red_text(ws, gen_row, [2, 3, 4])
                apply_red_text(ws, index, [9, 10])
                alternate_row = (index % 3) + 1
                apply_alternating_color(ws, gen_row, [1], alternate_row)
                apply_alternating_color(ws, index, [8], alternate_row)
                if discrepancies is not None:
                    _append_wrong_hours(
                        discrepancies,
                        store,
                        generated_data_map[employee_key],
                        webapp_records[-1],
                    )

    for generated_record in generated_records:
        if generated_record["matched"]:
            continue
        for webapp_record in webapp_records:
            if webapp_record["matched"]:
                continue
            if not _has_same_hours(generated_record, webapp_record):
                continue

            generated_record["matched"] = True
            webapp_record["matched"] = True
            if _has_missing_employee_pair(generated_record, webapp_record):
                apply_red_text(ws, generated_record["row"], [1])
                apply_red_text(ws, webapp_record["row"], [8])
            if discrepancies is not None and _has_missing_employee_pair(generated_record, webapp_record):
                _append_missing_employee_number(discrepancies, store, generated_record, webapp_record)
            break

    while True:
        unmatched_missing_generated = [
            record for record in generated_records
            if not record["matched"] and _is_missing_employee_number(record["employee_number"])
        ]
        unmatched_numbered_generated = [
            record for record in generated_records
            if not record["matched"] and not _is_missing_employee_number(record["employee_number"])
        ]
        unmatched_missing_webapp = [
            record for record in webapp_records
            if not record["matched"] and _is_missing_employee_number(record["employee_number"])
        ]
        unmatched_numbered_webapp = [
            record for record in webapp_records
            if not record["matched"] and not _is_missing_employee_number(record["employee_number"])
        ]

        paired = False
        if len(unmatched_numbered_generated) == 1 and len(unmatched_missing_webapp) == 1:
            generated_record = unmatched_numbered_generated[0]
            webapp_record = unmatched_missing_webapp[0]
            paired = True

        if not paired:
            break

        generated_record["matched"] = True
        webapp_record["matched"] = True
        has_hours_discrepancy = True
        apply_red_text(ws, generated_record["row"], [1, 2, 3, 4])
        apply_red_text(ws, webapp_record["row"], [8, 9, 10])
        if discrepancies is not None:
            _append_missing_employee_number(discrepancies, store, generated_record, webapp_record)
            _append_wrong_hours(discrepancies, store, generated_record, webapp_record)

    generated_only_rows = {
        record["row"]
        for record in generated_records
        if not record["matched"]
    }
    if generated_only_rows:
        has_hours_discrepancy = True
    for index, row in enumerate(generated_sorted.iterrows(), start=5):
        if index in generated_only_rows:
            ws.cell(row=index, column=1).fill = ORANGE_FILL
            if discrepancies is not None:
                discrepancies["wrong_hours"].append({
                    "store": store,
                    "employee_number": _employee_number(row[1]["Employee Number"]),
                    "employee_name": row[1]["Employee Name"],
                    "centech_regular_hours": _to_float(row[1]["Regular Hours"]),
                    "centech_overtime_hours": _to_float(row[1]["Overtime Hours"]),
                    "qa_regular_hours": None,
                    "qa_overtime_hours": None,
                })

    webapp_only_rows = {
        record["row"]
        for record in webapp_records
        if not record["matched"]
    }
    if webapp_only_rows:
        has_hours_discrepancy = True
    for index, row in enumerate(webapp_sorted.iterrows(), start=5):
        if index in webapp_only_rows:
            ws.cell(row=index, column=8).fill = YELLOW_FILL
            if discrepancies is not None:
                discrepancies["wrong_hours"].append({
                    "store": store,
                    "employee_number": _employee_number(row[1]["Employee Number"]),
                    "employee_name": None,
                    "centech_regular_hours": None,
                    "centech_overtime_hours": None,
                    "qa_regular_hours": _to_float(row[1]["Regular Hours"]),
                    "qa_overtime_hours": _to_float(row[1]["Overtime Hours"]),
                })

    if centech_tips and store in centech_tips:
        total_tip_generated = centech_tips[store]

    tip_rate_generated = total_tip_generated / total_hours_generated if total_hours_generated > 0 else 0
    tip_rate_webapp = total_tip_webapp / total_hours_webapp if total_hours_webapp > 0 else 0

    last_row = max(len(generated_sorted), len(webapp_sorted)) + 7

    ws.cell(row=last_row, column=1, value="Total Hours")
    ws.cell(row=last_row, column=2, value=total_hours_generated)
    ws.cell(row=last_row, column=8, value="Total Hours")
    ws.cell(row=last_row, column=9, value=total_hours_webapp)

    ws.cell(row=last_row + 1, column=1, value="Total Tips")
    ws.cell(row=last_row + 1, column=2, value=total_tip_generated)
    ws.cell(row=last_row + 1, column=8, value="Total Tips")
    ws.cell(row=last_row + 1, column=9, value=total_tip_webapp)

    ws.cell(row=last_row + 2, column=1, value="Tip Rate")
    ws.cell(row=last_row + 2, column=2, value=tip_rate_generated)
    ws.cell(row=last_row + 2, column=8, value="Tip Rate")
    ws.cell(row=last_row + 2, column=9, value=tip_rate_webapp)

    style_summary_row(ws, start_row=last_row, columns=[1, 8])
    style_summary_row(ws, start_row=last_row + 1, columns=[1, 8])
    style_summary_row(ws, start_row=last_row + 2, columns=[1, 8])

    if round(total_hours_generated, 2) != round(total_hours_webapp, 2):
        has_hours_discrepancy = True
        apply_red_text(ws, last_row, [2, 9])

    if round(total_tip_generated, 2) != round(total_tip_webapp, 2):
        has_tips_discrepancy = True
        apply_red_text(ws, last_row + 1, [2, 9])
        if discrepancies is not None:
            discrepancies["wrong_tips"].append({
                "store": store,
                "date": tip_date_label or "",
            })

    if has_hours_discrepancy:
        ws.sheet_properties.tabColor = "FF0000"
    elif has_tips_discrepancy:
        ws.sheet_properties.tabColor = "FFFF00"

    autofit_columns(ws)


def save_workbook(wb, output_path: Path) -> None:
    wb.save(output_path)
