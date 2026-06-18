from __future__ import annotations

from dataclasses import dataclass
from datetime import date, timedelta
from pathlib import Path

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

if __package__ in {None, ""}:
    import sys

    sys.path.insert(0, str(Path(__file__).resolve().parent))
    import common  # type: ignore
else:
    from . import common

CATEGORY_NAME = "Third Party Surcharge"
OUTPUT_SLUG = "third_party_surcharge"
LOOKBACK_DAYS = 10
LOOKAHEAD_DAYS = 30

PROVIDER_NAME_IDS = {
    "UberEats": {"4001"},
    "DoorDash": {"4004"},
    "GrubHub": {"4003"},
    "Postmate": {"4052"},
    "Amazon": {"amazon"},
}
PROVIDER_ORDER = [
    "UberEats",
    "DoorDash",
    "GrubHub",
    "Postmate",
    "Amazon",
    "Other",
    "House Account",
    "Total",
]
PROVIDER_METRICS = ["Ticket Count", "Net Sales", "Surcharge Sales", "Sales Tax", "Total Sales"]
HOUSE_METRICS = ["Ticket Count", "Net Sales", "Sales Tax", "Total Sales"]


@dataclass(frozen=True)
class DateRange:
    start_date: date
    end_date: date

    @property
    def start_date_str(self) -> str:
        return self.start_date.isoformat()

    @property
    def end_date_str(self) -> str:
        return self.end_date.isoformat()

    @property
    def period_key(self) -> str:
        return f"{self.start_date_str}_{self.end_date_str}"

    @property
    def target_dates(self) -> list[date]:
        return _date_list(self.start_date, self.end_date)

    @property
    def scan_dates(self) -> list[date]:
        return _date_list(
            self.start_date - timedelta(days=LOOKBACK_DAYS),
            self.end_date + timedelta(days=LOOKAHEAD_DAYS),
        )


def _date_list(start: date, end: date) -> list[date]:
    out: list[date] = []
    current = start
    while current <= end:
        out.append(current)
        current += timedelta(days=1)
    return out


def _parse_date(raw: str) -> date:
    try:
        return date.fromisoformat(raw.strip())
    except ValueError as exc:
        raise ValueError(f"Use YYYY-MM-DD format, got: {raw!r}") from exc


def _prompt_date(label: str) -> date:
    while True:
        raw = input(label).strip()
        if not raw:
            print("Input required.")
            continue
        try:
            return _parse_date(raw)
        except ValueError as exc:
            print(exc)


def _prompt_store_number() -> str:
    while True:
        raw = input("Store number: ").strip()
        if raw:
            return raw
        print("Input required.")


def _prompt_inputs(pos_data_dir: Path) -> tuple[DateRange, str]:
    print("=== Third Party Surcharge Audit ===")
    available_dates = common.list_available_dates(pos_data_dir)
    if available_dates:
        print(
            f"Available dates: {available_dates[0]} -> {available_dates[-1]} "
            f"({len(available_dates)} days)"
        )
    start_date = _prompt_date("Start date (YYYY-MM-DD): ")
    end_date = _prompt_date("End date (YYYY-MM-DD): ")
    if start_date > end_date:
        raise SystemExit("Start date must be <= end date.")

    date_range = DateRange(start_date=start_date, end_date=end_date)
    print(
        "Scan window: "
        f"{date_range.scan_dates[0].isoformat()} -> {date_range.scan_dates[-1].isoformat()} "
        f"({LOOKBACK_DAYS}-day backcheck, {LOOKAHEAD_DAYS}-day forward check)"
    )
    return date_range, _prompt_store_number()


def _load_store_id(pos_data_dir: Path, target_date_str: str, store_number: str) -> str:
    return common.load_store_id(pos_data_dir / target_date_str, store_number)


def _normalize_payments(pay: pd.DataFrame) -> pd.DataFrame:
    if pay.empty:
        return pay
    pay = pay.copy()
    for col in ["Payment_Type_ID", "Payment_Name_ID", "Name", "Ticket_Number"]:
        if col in pay.columns:
            pay[col] = pay[col].astype(str).str.strip()
    return pay


def _normalize_sts(sts: pd.DataFrame) -> pd.DataFrame:
    sts = sts.copy()
    for col in ["Taxable_Amount", "Non_Taxable_Amount", "Total"]:
        if col in sts.columns:
            sts[col] = common.to_num(sts[col])
    if "Ticket_Number" in sts.columns:
        sts["Ticket_Number"] = sts["Ticket_Number"].astype(str)
    if "Category_ID" in sts.columns:
        sts["Category_ID"] = sts["Category_ID"].astype(str).str.strip()
    return sts


def _sts_total(sts: pd.DataFrame, ticket_set: set[str], category_id: str) -> float:
    if sts.empty or not ticket_set:
        return 0.0
    mask = sts["Ticket_Number"].astype(str).isin(ticket_set) & sts["Category_ID"].eq(category_id)
    return float(sts.loc[mask, "Total"].sum())


def _bucket_metrics(sts: pd.DataFrame, ticket_set: set[str], *, include_surcharge: bool) -> dict[str, float | int]:
    net_sales = _sts_total(sts, ticket_set, "4")
    surcharge_sales = _sts_total(sts, ticket_set, "11") if include_surcharge else 0.0
    sales_tax = _sts_total(sts, ticket_set, "5")
    total_sales = net_sales + surcharge_sales + sales_tax
    out: dict[str, float | int] = {
        "Ticket Count": len(ticket_set),
        "Net Sales": round(net_sales, 2),
        "Sales Tax": round(sales_tax, 2),
        "Total Sales": round(total_sales, 2),
    }
    if include_surcharge:
        out["Surcharge Sales"] = round(surcharge_sales, 2)
    return out


def _provider_ticket_sets(pay_store: pd.DataFrame) -> dict[str, set[str]]:
    out: dict[str, set[str]] = {name: set() for name in PROVIDER_ORDER if name != "Total"}
    if pay_store.empty:
        return out

    type_13 = pay_store[pay_store["Payment_Type_ID"].eq("13")].copy()
    type_13["Name"] = type_13["Name"].astype(str).str.strip()
    mapped_names = set().union(*PROVIDER_NAME_IDS.values())
    for provider, names in PROVIDER_NAME_IDS.items():
        provider_names = {name.lower() for name in names}
        out[provider] = set(
            type_13[type_13["Name"].str.lower().isin(provider_names)]["Ticket_Number"].astype(str)
        )

    mapped_lower = {name.lower() for name in mapped_names}
    out["Other"] = set(
        type_13[~type_13["Name"].str.lower().isin(mapped_lower)]["Ticket_Number"].astype(str)
    )
    out["House Account"] = set(
        pay_store[pay_store["Payment_Type_ID"].eq("7")]["Ticket_Number"].astype(str)
    )
    return out


def _compute_store_day(
    *,
    pos_data_dir: Path,
    target_date_str: str,
    store_number: str,
    cross_date_sales: dict[str, tuple[pd.DataFrame, pd.DataFrame]],
    cross_date_payments: dict[str, pd.DataFrame],
) -> tuple[dict[tuple[str, str], float | int | str], dict[str, object]]:
    store_id = _load_store_id(pos_data_dir, target_date_str, store_number)
    if target_date_str not in cross_date_sales:
        raise ValueError(f"No cross-date sales context found for {target_date_str}")

    st, sts = cross_date_sales[target_date_str]
    store_st = st[st["Store_ID"].astype(str).str.strip().eq(str(store_id))].copy()
    store_tix = set(store_st["Ticket_Number"].astype(str))
    sts = _normalize_sts(sts)

    pay = _normalize_payments(cross_date_payments.get(target_date_str, pd.DataFrame()))
    pay_store = pay[pay["Ticket_Number"].astype(str).isin(store_tix)].copy() if not pay.empty else pay
    store_paid = store_tix & set(pay_store["Ticket_Number"].astype(str)) if not pay_store.empty else set()

    bucket_sets = _provider_ticket_sets(pay_store)
    for provider in bucket_sets:
        bucket_sets[provider] = bucket_sets[provider] & store_paid

    row: dict[tuple[str, str], float | int | str] = {("", "Store Number"): store_number}
    detail_row: dict[str, object] = {
        "date": target_date_str,
        "store_number": store_number,
        "store_id": store_id,
        "store_paid_tickets": len(store_paid),
    }

    total_ticket_set: set[str] = set()
    for provider in PROVIDER_ORDER:
        if provider == "Total":
            continue
        total_ticket_set |= bucket_sets[provider]
        include_surcharge = provider != "House Account"
        metrics = _bucket_metrics(sts, bucket_sets[provider], include_surcharge=include_surcharge)
        metric_order = PROVIDER_METRICS if include_surcharge else HOUSE_METRICS
        for metric in metric_order:
            row[(provider, metric)] = metrics[metric]
            detail_row[f"{provider.lower().replace(' ', '_')}_{metric.lower().replace(' ', '_')}"] = metrics[metric]

    total_metrics = _bucket_metrics(sts, total_ticket_set, include_surcharge=True)
    for metric in PROVIDER_METRICS:
        row[("Total", metric)] = total_metrics[metric]
        detail_row[f"total_{metric.lower().replace(' ', '_')}"] = total_metrics[metric]

    multi_bucket_count = sum(
        count > 1
        for count in pd.Series(
            [ticket for tickets in bucket_sets.values() for ticket in tickets], dtype=str
        ).value_counts()
    )
    detail_row["multi_bucket_ticket_count"] = int(multi_bucket_count)
    return row, detail_row


def _sum_multiindex_rows(rows: list[dict[tuple[str, str], float | int | str]], store_number: str) -> pd.DataFrame:
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows)
    numeric_cols = [col for col in df.columns if col != ("", "Store Number")]
    summed: dict[tuple[str, str], float | int | str] = {("", "Store Number"): store_number}
    for col in numeric_cols:
        summed[col] = round(float(pd.to_numeric(df[col], errors="coerce").fillna(0).sum()), 2)
        if col[1] == "Ticket Count":
            summed[col] = int(summed[col])
    out = pd.DataFrame([summed])
    out.columns = pd.MultiIndex.from_tuples(out.columns)
    return out


def _write_outputs(
    out_dir: Path,
    report_df: pd.DataFrame,
    daily_report_df: pd.DataFrame,
    detail_df: pd.DataFrame,
    summary_df: pd.DataFrame,
) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    report_df.to_csv(out_dir / "third_party_surcharge_report.csv", index=False)
    daily_report_df.to_csv(out_dir / "third_party_surcharge_by_date.csv", index=False)
    detail_df.to_csv(out_dir / "audit_third_party_surcharge_detail.csv", index=False)
    summary_df.to_csv(out_dir / "audit_summary.csv", index=False)

    with pd.ExcelWriter(out_dir / "audit_third_party_surcharge.xlsx", engine="openpyxl") as writer:
        _write_multiindex_sheet(writer, "report", report_df)
        _write_multiindex_sheet(writer, "by_date", daily_report_df)
        summary_df.to_excel(writer, sheet_name="summary", index=False)
        detail_df.to_excel(writer, sheet_name="detail", index=False)


def _write_multiindex_sheet(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame) -> None:
    ws = writer.book.create_sheet(sheet_name)
    columns = list(df.columns)
    ws.append([str(col[0]) if col[0] else "" for col in columns])
    ws.append([str(col[1]) for col in columns])
    for row in df.itertuples(index=False, name=None):
        ws.append(list(row))

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    start_col = 1
    current = ws.cell(row=1, column=1).value
    for col_idx in range(2, ws.max_column + 2):
        value = ws.cell(row=1, column=col_idx).value if col_idx <= ws.max_column else None
        if value != current:
            if current and col_idx - start_col > 1:
                ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=col_idx - 1)
            start_col = col_idx
            current = value

    ws.freeze_panes = "A3"
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for cell in col_cells:
            max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 18)


def run() -> None:
    root = common.repo_root()
    pos_data_dir = root / "pos_data"
    date_range, store_number = _prompt_inputs(pos_data_dir)

    existing_scan_dates = [
        d for d in date_range.scan_dates if (pos_data_dir / d.isoformat()).is_dir()
    ]
    cross_date_sales = common._build_cross_date_sales_context(  # type: ignore[attr-defined]
        pos_data_dir, existing_scan_dates
    )
    cross_date_payments = common._build_cross_date_payments(  # type: ignore[attr-defined]
        pos_data_dir, existing_scan_dates
    )

    report_rows: list[dict[tuple[str, str], float | int | str]] = []
    detail_rows: list[dict[str, object]] = []
    missing_dates: list[str] = []
    skipped: list[dict[str, str]] = []

    for target_date in date_range.target_dates:
        target_date_str = target_date.isoformat()
        if not (pos_data_dir / target_date_str).is_dir():
            missing_dates.append(target_date_str)
            continue
        try:
            report_row, detail_row = _compute_store_day(
                pos_data_dir=pos_data_dir,
                target_date_str=target_date_str,
                store_number=store_number,
                cross_date_sales=cross_date_sales,
                cross_date_payments=cross_date_payments,
            )
            report_rows.append(report_row)
            detail_rows.append(detail_row)
        except Exception as exc:
            skipped.append(
                {"date": target_date_str, "store_number": store_number, "reason": str(exc)}
            )

    if not report_rows:
        raise SystemExit("No third-party surcharge rows were produced.")

    daily_report_df = pd.DataFrame(report_rows)
    daily_report_df.columns = pd.MultiIndex.from_tuples(daily_report_df.columns)
    report_df = _sum_multiindex_rows(report_rows, store_number)
    detail_df = pd.DataFrame(detail_rows)

    summary_rows: list[dict[str, object]] = [
        {"metric": "start_date", "value": date_range.start_date_str},
        {"metric": "end_date", "value": date_range.end_date_str},
        {"metric": "scan_start_date", "value": date_range.scan_dates[0].isoformat()},
        {"metric": "scan_end_date", "value": date_range.scan_dates[-1].isoformat()},
        {"metric": "store_number", "value": store_number},
        {"metric": "store_day_rows", "value": len(detail_df)},
        {"metric": "missing_date_folders", "value": len(missing_dates)},
        {"metric": "skipped_store_days", "value": len(skipped)},
        {"metric": "name_id_ubereats", "value": "4001"},
        {"metric": "name_id_doordash", "value": "4004"},
        {"metric": "name_id_grubhub", "value": "4003"},
        {"metric": "name_id_postmate", "value": "4052"},
        {"metric": "amazon_mapping", "value": "no numeric POS Name ID found; matches literal 'amazon' if present"},
        {"metric": "other_mapping", "value": "Payment_Type_ID 13 names not mapped above"},
        {"metric": "house_account_mapping", "value": "Payment_Type_ID 7"},
    ]
    if missing_dates:
        summary_rows.append({"metric": "missing_dates", "value": ",".join(missing_dates)})
    summary_df = pd.DataFrame(summary_rows)

    out_dir = (
        common.script_root()
        / "audits"
        / OUTPUT_SLUG
        / date_range.period_key
        / str(store_number)
    )
    _write_outputs(out_dir, report_df, daily_report_df, detail_df, summary_df)

    if skipped:
        pd.DataFrame(skipped).to_csv(out_dir / "audit_third_party_surcharge_skipped.csv", index=False)

    print(f"Audit written to: {out_dir}")


if __name__ == "__main__":
    run()
