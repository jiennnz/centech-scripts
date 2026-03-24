from __future__ import annotations

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

color_1 = "DAF2D0"
color_2 = "F2CEEF"
color_3 = "FBE2D5"


def autofit_columns(ws) -> None:
    employee_number_col = get_column_letter(1)
    highlight_col = get_column_letter(14)
    ws.column_dimensions[employee_number_col].width = 20
    ws.column_dimensions[highlight_col].width = 10

    for col in range(1, ws.max_column + 1):
        column = get_column_letter(col)
        if column in [employee_number_col, highlight_col]:
            continue
        max_length = 0
        for row in ws.iter_rows(min_col=col, max_col=col, max_row=ws.max_row):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2


def style_summary_row(ws, start_row: int, columns: list) -> None:
    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    for col in columns:
        cell = ws.cell(row=start_row, column=col)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)
        cell.fill = blue_fill


def apply_red_text(ws, row: int, columns: list) -> None:
    red_font = Font(color="FF0000", bold=True)
    for col in columns:
        ws.cell(row=row, column=col).font = red_font


def apply_alternating_color(ws, row: int, columns: list, alternate_row: int) -> None:
    color = {1: color_1, 2: color_2}.get(alternate_row, color_3)
    for col in columns:
        ws.cell(row=row, column=col).fill = PatternFill(
            start_color=color, end_color=color, fill_type="solid"
        )


def style_sheet(ws) -> None:
    ws.merge_cells("A1:L1")
    main_title = ws["A1"]
    main_title.value = "Tips Computation Comparison"
    main_title.alignment = Alignment(horizontal="center", vertical="center")
    main_title.font = Font(size=18, bold=True, color="FFFFFF")
    main_title.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

    ws.merge_cells("A2:F3")
    century_title = ws["A2"]
    century_title.value = "QA"
    century_title.alignment = Alignment(horizontal="center", vertical="center")
    century_title.font = Font(size=20, bold=True)
    century_title.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

    century_headers = ["Employee Number", "Employee Name", "Regular Hours", "Overtime Hours", "Coded Amount", "Tip Rate"]
    for col_num, header in enumerate(century_headers, start=1):
        cell = ws.cell(row=4, column=col_num, value=header)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    ws.merge_cells("H2:L3")
    qsr_title = ws["H2"]
    qsr_title.value = "Century"
    qsr_title.alignment = Alignment(horizontal="center", vertical="center")
    qsr_title.font = Font(size=20, bold=True)
    qsr_title.fill = PatternFill(start_color="CAEDFB", end_color="CAEDFB", fill_type="solid")

    qsr_headers = ["Employee Number", "Regular Hours", "Overtime Hours", "Coded Amount", "Tip Rate"]
    for col_num, header in enumerate(qsr_headers, start=8):
        cell = ws.cell(row=4, column=col_num, value=header)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    ws.merge_cells("N2:O2")
    highlighted_cells_title = ws["N2"]
    highlighted_cells_title.value = "Highlighted Cells Guide"
    highlighted_cells_title.alignment = Alignment(horizontal="center", vertical="center")
    highlighted_cells_title.font = Font(size=14, bold=True, color="FFFFFF")
    highlighted_cells_title.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

    ws["N3"].fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    ws["O3"].value = "Only in Century"
    ws["O3"].alignment = Alignment(horizontal="center", vertical="center")
    ws["N4"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws["O4"].value = "Only in QSR"
    ws["O4"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("N6:O10")
    ws["N6"].value = "Employee Names highlighted in Red have missing clock in/out. The cell colors beside them are only guidelines for the correct record from the QSR table"
    ws["N6"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["N6"].font = Font(size=12, color="000000")