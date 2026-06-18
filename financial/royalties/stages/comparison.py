from __future__ import annotations

import csv
import re
from collections.abc import Sequence
from dataclasses import dataclass
from datetime import date
from pathlib import Path

import pandas as pd
from dateutil import parser as date_parser
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from financial.sales_export_comparison.stages.template_builder import (
    TableLayout,
    _append_shifted_conditional_formatting,
    _copy_table_block,
    _snapshot_conditional_formatting,
    _stamp_store_headers,
)
from financial.sales_export_comparison.stages.heatmap import (
    HeatmapConfig,
    run as run_heatmap,
)


ROYALTY_LAYOUT = TableLayout(
    start_row=1,
    end_row=11,
    start_col=2,
    end_col=5,
    spacer_cols=1,
    category_col=1,
    data_start_row=4,
    data_end_row=9,
)

REQUIRED_COLUMNS = (
    "DateRange",
    "Class",
    "Transaction Category",
    "Account Number",
    "Account Name",
    "Debit",
    "Credit",
)

ACCOUNT_MAPPING_SHEET = "Account Mapping Check"

_HEADER_FILL = PatternFill(fill_type="solid", start_color="1F4E79", end_color="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_WARN_FILL = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")
_OK_FILL = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")


@dataclass(frozen=True, order=True)
class RoyaltyPeriod:
    start: date
    end: date

    @property
    def key(self) -> str:
        return f"{self.start.isoformat()}_{self.end.isoformat()}"

    @property
    def sheet_title(self) -> str:
        if self.start == self.end:
            return self.start.strftime("%b %d")
        return f"{self.start:%b %d}-{self.end:%b %d}"


@dataclass(frozen=True)
class RoyaltyComparisonConfig:
    template_path: Path
    output_path: Path
    centech_path: Path | Sequence[Path]
    client_path: Path | Sequence[Path]
    stores: list[str]
    start_date: date
    end_date: date
    source_label: str = "Client"
    centech_label: str = "CenTech"
    tolerance: float = 0.01


@dataclass(frozen=True)
class RoyaltyComparisonResult:
    output_path: Path
    periods: tuple[RoyaltyPeriod, ...]
    centech_rows_written: int
    client_rows_written: int
    account_mapping_differences: int


def _read_export(path: Path) -> pd.DataFrame:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xls"}:
        df = pd.read_excel(path, engine="openpyxl")
    elif suffix == ".csv":
        df = _read_csv_export(path)
    else:
        raise ValueError(f"Unsupported export format: {path}")

    df = _normalize_export_columns(df)
    missing = [col for col in REQUIRED_COLUMNS if col not in df.columns]
    if missing:
        raise ValueError(f"{path.name} is missing required columns: {', '.join(missing)}")
    return df


def _read_exports(paths: Path | Sequence[Path]) -> pd.DataFrame:
    if isinstance(paths, Path):
        normalized_paths = (paths,)
    else:
        normalized_paths = tuple(Path(path) for path in paths)

    if not normalized_paths:
        raise ValueError("No royalty export paths provided.")

    frames = [_read_export(path) for path in normalized_paths]
    if len(frames) == 1:
        return frames[0]
    return pd.concat(frames, ignore_index=True)


def _read_csv_export(path: Path) -> pd.DataFrame:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        try:
            header = next(reader)
        except StopIteration as exc:
            raise ValueError(f"{path.name} is empty.") from exc

        expected_columns = len(header)
        non_empty_overflow: list[tuple[int, int]] = []
        for line_no, row in enumerate(reader, start=2):
            if len(row) <= expected_columns:
                continue
            extra_values = row[expected_columns:]
            if any(str(value).strip() for value in extra_values):
                non_empty_overflow.append((line_no, len(row)))

        if non_empty_overflow:
            details = ", ".join(
                f"line {line_no} has {field_count} fields"
                for line_no, field_count in non_empty_overflow[:5]
            )
            raise ValueError(
                f"{path.name} has rows wider than its {expected_columns}-column header "
                f"with non-empty extra fields: {details}"
            )

    return pd.read_csv(path, encoding="utf-8-sig", usecols=range(expected_columns))


def _normalize_export_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.columns = [str(col).strip() for col in out.columns]
    if "DateRange" not in out.columns and "Date" in out.columns:
        out = out.rename(columns={"Date": "DateRange"})
    return out


def _parse_period(raw: object) -> RoyaltyPeriod | None:
    if pd.isna(raw):
        return None
    text = str(raw).strip()
    parts = re.split(r"\s+-\s+", text, maxsplit=1)
    try:
        start = date_parser.parse(parts[0]).date()
        end = date_parser.parse(parts[1]).date() if len(parts) == 2 else start
    except (TypeError, ValueError, OverflowError):
        return None
    if start > end:
        start, end = end, start
    return RoyaltyPeriod(start=start, end=end)


def _period_sort_key(period: RoyaltyPeriod) -> tuple[date, int, date]:
    duration_days = (period.end - period.start).days
    return (period.start, -duration_days, period.end)


def _filter_frame(df: pd.DataFrame, *, stores: set[str], start_date: date, end_date: date) -> pd.DataFrame:
    out = df.copy()
    out["_period"] = out["DateRange"].map(_parse_period)
    out["_store"] = out["Class"].map(_normalize_store)
    out["_category"] = out["Transaction Category"].map(_clean_text)
    out["Debit"] = pd.to_numeric(out["Debit"], errors="coerce").fillna(0.0)
    out["Credit"] = pd.to_numeric(out["Credit"], errors="coerce").fillna(0.0)

    period_mask = out["_period"].map(
        lambda p: p is not None and p.start <= end_date and p.end >= start_date
    )
    store_mask = out["_store"].isin(stores)
    summary_mask = out["_category"].str.upper().ne("SUMMARY")
    return out[period_mask & store_mask & summary_mask].copy()


def _normalize_store(raw: object) -> str:
    if pd.isna(raw):
        return ""
    text = str(raw).strip()
    if re.fullmatch(r"\d+\.0", text):
        return text[:-2]
    return text.split()[0].split("-")[0].strip()


def _clean_text(raw: object) -> str:
    if pd.isna(raw):
        return ""
    return str(raw).strip()


def _category_rows(ws) -> dict[str, int]:
    rows: dict[str, int] = {}
    for row_idx in range(ROYALTY_LAYOUT.data_start_row, ROYALTY_LAYOUT.data_end_row + 1):
        value = ws.cell(row_idx, ROYALTY_LAYOUT.category_col).value
        if isinstance(value, str) and value.strip():
            rows[value.strip()] = row_idx
    return rows


def _royalty_category_rows() -> dict[str, int]:
    return {
        "Royalty Fee": 4,
        "Royalties Bank Acct Entry": 5,
        "National Media Fee": 6,
        "National Media Bank Entry": 7,
        "Corporate Advertising Fee": 8,
        "Corp Advertising Bank Acct Entry": 9,
    }


def _validate_filtered_export(df: pd.DataFrame, *, label: str) -> None:
    if df.empty:
        raise ValueError(f"No {label} rows matched the selected date range and organization stores.")

    expected_categories = set(_royalty_category_rows())
    observed_categories = {category for category in df["_category"].dropna().unique() if category}
    if observed_categories & expected_categories:
        return

    expected = ", ".join(sorted(expected_categories))
    observed = ", ".join(sorted(observed_categories)[:12]) or "none"
    if len(observed_categories) > 12:
        observed += ", ..."
    raise ValueError(
        f"No {label} rows matched the expected royalty categories. "
        f"Expected one of: {expected}. Found categories: {observed}."
    )


def _build_workbook(
    *,
    template_path: Path,
    output_path: Path,
    periods: list[RoyaltyPeriod],
    stores: list[str],
    source_label: str,
    centech_label: str,
) -> None:
    if not periods:
        raise ValueError("No royalty DateRange values found for the selected period.")

    workbook = load_workbook(template_path)
    template_sheet = workbook[workbook.sheetnames[0]]
    template_cf = _snapshot_conditional_formatting(template_sheet)

    for period in periods:
        ws = workbook.copy_worksheet(template_sheet)
        ws.title = _safe_sheet_title(period.sheet_title, workbook.sheetnames)
        ws.freeze_panes = template_sheet.freeze_panes
        for idx in range(len(stores)):
            if idx > 0:
                _copy_table_block(ws, ROYALTY_LAYOUT, idx)
            _append_shifted_conditional_formatting(
                ws,
                template_cf,
                col_shift=idx * ROYALTY_LAYOUT.block_width,
            )
        _stamp_store_headers(
            ws,
            stores=stores,
            source_label=source_label,
            layout=ROYALTY_LAYOUT,
            centech_label=centech_label,
        )

    workbook.remove(template_sheet)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def _safe_sheet_title(title: str, existing: list[str]) -> str:
    clean = re.sub(r"[\[\]:*?/\\]", "-", title).strip()[:31] or "Period"
    if clean not in existing:
        return clean
    base = clean[:28]
    idx = 2
    while True:
        candidate = f"{base}-{idx}"[:31]
        if candidate not in existing:
            return candidate
        idx += 1


def _write_side(
    *,
    workbook_path: Path,
    df: pd.DataFrame,
    side: str,
    periods: list[RoyaltyPeriod],
    period_sheets: dict[RoyaltyPeriod, str],
    stores: list[str],
) -> int:
    wb = load_workbook(workbook_path)
    rows_written = 0

    grouped = (
        df.groupby(["_period", "_store", "_category"], dropna=False)
        .agg(Debit=("Debit", "sum"), Credit=("Credit", "sum"))
        .reset_index()
    )

    for period in periods:
        ws = wb[period_sheets[period]]
        category_rows = _category_rows(ws)
        store_columns = _store_columns(side=side, stores=stores)
        period_rows = grouped[grouped["_period"] == period]

        for _, row in period_rows.iterrows():
            category = row["_category"]
            store = row["_store"]
            row_idx = category_rows.get(category)
            columns = store_columns.get(store)
            if row_idx is None or columns is None:
                continue

            debit_col, credit_col = columns
            _add_numeric(ws.cell(row_idx, debit_col), row["Debit"])
            _add_numeric(ws.cell(row_idx, credit_col), row["Credit"])
            rows_written += 1

    wb.save(workbook_path)
    return rows_written


def _store_columns(*, side: str, stores: list[str]) -> dict[str, tuple[int, int]]:
    if side not in {"centech", "client"}:
        raise ValueError(f"Unknown side: {side}")
    offset = 0 if side == "centech" else 2
    cols: dict[str, tuple[int, int]] = {}
    for idx, store in enumerate(stores):
        start_col = ROYALTY_LAYOUT.start_col + (idx * ROYALTY_LAYOUT.block_width)
        cols[str(store)] = (start_col + offset, start_col + offset + 1)
    return cols


def _add_numeric(cell, amount: object) -> None:
    value = float(amount or 0.0)
    if value == 0:
        return
    existing = cell.value
    if isinstance(existing, str) and existing.startswith("="):
        return
    try:
        base = float(existing or 0)
    except (TypeError, ValueError):
        base = 0.0
    cell.value = round(base + value, 2)


def _format_account_number(raw: object) -> str:
    if pd.isna(raw) or str(raw).strip() == "":
        return ""
    try:
        value = float(raw)
    except (TypeError, ValueError):
        return str(raw).strip()
    if value.is_integer():
        return str(int(value))
    return str(raw).strip()


def _mapping_pairs(df: pd.DataFrame) -> dict[str, set[tuple[str, str]]]:
    pairs: dict[str, set[tuple[str, str]]] = {}
    for _, row in df.iterrows():
        category = _clean_text(row.get("_category"))
        if not category:
            continue
        pair = (
            _format_account_number(row.get("Account Number")),
            _clean_text(row.get("Account Name")),
        )
        pairs.setdefault(category, set()).add(pair)
    return pairs


def _join_pairs(pairs: set[tuple[str, str]], index: int) -> str:
    values = sorted({pair[index] for pair in pairs if pair[index]})
    return "; ".join(values)


def _mapping_differences(centech_df: pd.DataFrame, client_df: pd.DataFrame) -> list[dict[str, str]]:
    centech = _mapping_pairs(centech_df)
    client = _mapping_pairs(client_df)
    rows: list[dict[str, str]] = []

    for category in sorted(set(centech) | set(client)):
        c_pairs = centech.get(category, set())
        s_pairs = client.get(category, set())
        if c_pairs == s_pairs:
            continue

        issue = []
        if not c_pairs:
            issue.append("missing in CenTech")
        if not s_pairs:
            issue.append("missing in client")
        if c_pairs and s_pairs:
            if _join_pairs(c_pairs, 0) != _join_pairs(s_pairs, 0):
                issue.append("account number differs")
            if _join_pairs(c_pairs, 1) != _join_pairs(s_pairs, 1):
                issue.append("account name differs")

        rows.append(
            {
                "Transaction Category": category,
                "CenTech Account Number": _join_pairs(c_pairs, 0),
                "CenTech Account Name": _join_pairs(c_pairs, 1),
                "Client Account Number": _join_pairs(s_pairs, 0),
                "Client Account Name": _join_pairs(s_pairs, 1),
                "Issue": ", ".join(issue),
            }
        )
    return rows


def _write_account_mapping_sheet(
    *,
    workbook_path: Path,
    centech_df: pd.DataFrame,
    client_df: pd.DataFrame,
) -> int:
    wb = load_workbook(workbook_path)
    if ACCOUNT_MAPPING_SHEET in wb.sheetnames:
        del wb[ACCOUNT_MAPPING_SHEET]
    ws = wb.create_sheet(ACCOUNT_MAPPING_SHEET)

    rows = _mapping_differences(centech_df, client_df)
    headers = [
        "Transaction Category",
        "CenTech Account Number",
        "CenTech Account Name",
        "Client Account Number",
        "Client Account Name",
        "Issue",
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    if rows:
        for row_idx, item in enumerate(rows, start=2):
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=item[header])
                if header == "Issue":
                    cell.fill = _WARN_FILL
    else:
        cell = ws.cell(row=2, column=1, value="No account mapping differences found")
        cell.fill = _OK_FILL

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(len(rows) + 1, 2)}"
    _auto_width(ws)
    wb.save(workbook_path)
    return len(rows)


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 45)


def run(config: RoyaltyComparisonConfig) -> RoyaltyComparisonResult:
    stores = [str(store) for store in config.stores]
    store_set = set(stores)

    centech_df = _filter_frame(
        _read_exports(config.centech_path),
        stores=store_set,
        start_date=config.start_date,
        end_date=config.end_date,
    )
    client_df = _filter_frame(
        _read_exports(config.client_path),
        stores=store_set,
        start_date=config.start_date,
        end_date=config.end_date,
    )
    _validate_filtered_export(centech_df, label=config.centech_label)
    _validate_filtered_export(client_df, label=config.source_label)

    periods = sorted(
        set(centech_df["_period"].dropna()) | set(client_df["_period"].dropna()),
        key=_period_sort_key,
    )
    if not periods:
        raise ValueError(
            "No royalty rows matched the selected date range and organization stores."
        )

    _build_workbook(
        template_path=config.template_path,
        output_path=config.output_path,
        periods=periods,
        stores=stores,
        source_label=config.source_label,
        centech_label=config.centech_label,
    )

    wb = load_workbook(config.output_path, read_only=True)
    period_sheets = dict(zip(periods, wb.sheetnames, strict=True))
    wb.close()

    centech_rows = _write_side(
        workbook_path=config.output_path,
        df=centech_df,
        side="centech",
        periods=periods,
        period_sheets=period_sheets,
        stores=stores,
    )
    client_rows = _write_side(
        workbook_path=config.output_path,
        df=client_df,
        side="client",
        periods=periods,
        period_sheets=period_sheets,
        stores=stores,
    )

    run_heatmap(
        HeatmapConfig(
            workbook_path=config.output_path,
            stores=stores,
            category_rows=_royalty_category_rows(),
            tolerance=config.tolerance,
            layout=ROYALTY_LAYOUT,
            source_label=config.source_label,
            centech_label=config.centech_label,
        )
    )

    mapping_diff_count = _write_account_mapping_sheet(
        workbook_path=config.output_path,
        centech_df=centech_df,
        client_df=client_df,
    )

    return RoyaltyComparisonResult(
        output_path=config.output_path,
        periods=tuple(periods),
        centech_rows_written=centech_rows,
        client_rows_written=client_rows,
        account_mapping_differences=mapping_diff_count,
    )
