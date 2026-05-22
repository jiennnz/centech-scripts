from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from financial.sales_export_comparison.rules import OrgRule

SHEET_NAME = "Diagnostics"

_HEADER_FILL = PatternFill(fill_type="solid", start_color="1F4E79", end_color="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_SECTION_FILL = PatternFill(fill_type="solid", start_color="D6E4F0", end_color="D6E4F0")
_SECTION_FONT = Font(bold=True, color="1F4E79")
_WARN_FILL = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")
_OK_FILL = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")


@dataclass(frozen=True)
class DiagnosticsConfig:
    workbook_path: Path
    centech_path: Path
    org_rule: OrgRule
    start_date: date
    end_date: date
    tolerance: float = 0.01


def _read_centech(path: Path, rule: OrgRule) -> pd.DataFrame:
    suffix = path.suffix.lower()
    cfg = rule.centech
    if suffix in {".xlsx", ".xls"}:
        return pd.read_excel(path, engine="openpyxl", **getattr(cfg, "read_excel_kwargs", {}))
    kwargs = dict(getattr(cfg, "read_csv_kwargs", {}))
    return pd.read_csv(path, **kwargs)


def _apply_header_row(ws, row: int, cols: list[str]) -> None:
    for c, label in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _section_title(ws, row: int, title: str, ncols: int) -> None:
    cell = ws.cell(row=row, column=1, value=title)
    cell.fill = _SECTION_FILL
    cell.font = _SECTION_FONT
    for c in range(2, ncols + 1):
        ws.cell(row=row, column=c).fill = _SECTION_FILL


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 60)


def run(config: DiagnosticsConfig) -> None:
    rule = config.org_rule
    centech_cfg = rule.centech

    df = _read_centech(config.centech_path, rule)

    # Drop summary rows before any analysis
    if centech_cfg.skip_category:
        df = df[df[centech_cfg.category_column].astype(str).str.strip() != centech_cfg.skip_category]

    # Filter to period
    dates = pd.to_datetime(
        df[centech_cfg.date_column],
        format=getattr(centech_cfg, "date_parse_format", None) or None,
        errors="coerce",
    )
    mask = (dates.dt.date >= config.start_date) & (dates.dt.date <= config.end_date)
    df = df[mask].copy()
    df["_date"] = dates[mask].dt.date

    # Normalize numeric columns
    df[centech_cfg.debit_column] = pd.to_numeric(df[centech_cfg.debit_column], errors="coerce").fillna(0.0)
    df[centech_cfg.credit_column] = pd.to_numeric(df[centech_cfg.credit_column], errors="coerce").fillna(0.0)

    store_col = centech_cfg.store_column

    # --- Unbalanced stores ---
    grouped = df.groupby(["_date", store_col]).agg(
        total_debit=(centech_cfg.debit_column, "sum"),
        total_credit=(centech_cfg.credit_column, "sum"),
    ).reset_index()
    grouped["difference"] = grouped["total_debit"] - grouped["total_credit"]
    unbalanced = grouped[grouped["difference"].abs() > config.tolerance].copy()
    balanced = grouped[grouped["difference"].abs() <= config.tolerance].copy()

    # --- Missing account mapping ---
    mapping_cols = ["Account Number", "Account Name", "Memo"]
    present_cols = [c for c in mapping_cols if c in df.columns]

    missing_rows: list[dict] = []
    for _, row in df.iterrows():
        missing_fields = [c for c in present_cols if pd.isna(row[c]) or str(row[c]).strip() == ""]
        if missing_fields:
            missing_rows.append({
                "Date": row["_date"],
                "Store": row[store_col],
                "Category": row.get(centech_cfg.category_column, ""),
                "Missing Fields": ", ".join(missing_fields),
            })

    # --- Write to workbook ---
    wb = load_workbook(config.workbook_path)

    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME)

    cursor = 1

    # Section 1: Unbalanced Stores
    _section_title(ws, cursor, "Unbalanced Stores", 5)
    cursor += 1
    _apply_header_row(ws, cursor, ["Date", "Store", "Total Debit", "Total Credit", "Difference"])
    cursor += 1

    if unbalanced.empty:
        cell = ws.cell(row=cursor, column=1, value="All stores balanced")
        cell.fill = _OK_FILL
        cursor += 1
    else:
        for _, r in unbalanced.iterrows():
            ws.cell(row=cursor, column=1, value=str(r["_date"]))
            ws.cell(row=cursor, column=2, value=r[store_col])
            ws.cell(row=cursor, column=3, value=round(r["total_debit"], 2))
            ws.cell(row=cursor, column=4, value=round(r["total_credit"], 2))
            diff_cell = ws.cell(row=cursor, column=5, value=round(r["difference"], 2))
            diff_cell.fill = _WARN_FILL
            cursor += 1

    cursor += 1  # spacer

    # Section 2: Missing Account Mapping
    _section_title(ws, cursor, "Missing Account Mapping (Account Number / Account Name / Memo)", 4)
    cursor += 1

    if not present_cols:
        ws.cell(row=cursor, column=1, value="No mapping columns found in export")
        cursor += 1
    elif not missing_rows:
        cell = ws.cell(row=cursor, column=1, value="No missing mappings found")
        cell.fill = _OK_FILL
        cursor += 1
    else:
        _apply_header_row(ws, cursor, ["Date", "Store", "Category", "Missing Fields"])
        cursor += 1
        for mr in missing_rows:
            ws.cell(row=cursor, column=1, value=str(mr["Date"]))
            ws.cell(row=cursor, column=2, value=mr["Store"])
            ws.cell(row=cursor, column=3, value=mr["Category"])
            miss_cell = ws.cell(row=cursor, column=4, value=mr["Missing Fields"])
            miss_cell.fill = _WARN_FILL
            cursor += 1

    _auto_width(ws)
    wb.save(config.workbook_path)
