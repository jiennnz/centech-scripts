from __future__ import annotations

import re
from copy import copy, deepcopy
from dataclasses import dataclass
from datetime import date, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.cell_range import CellRange, MultiCellRange
from openpyxl.worksheet.worksheet import Worksheet


# Match A1-style refs inside CF formula strings (not sheet-qualified).
_CF_CELL_REF_RE = re.compile(
    r"(?<![A-Za-z0-9_])(\$?)([A-Za-z]{1,3})(\$?)(\d+)(?![A-Za-z0-9_])"
)


@dataclass(frozen=True)
class TableLayout:
    start_row: int = 1
    end_row: int = 30
    start_col: int = 2  # B
    end_col: int = 5  # E
    spacer_cols: int = 1
    category_col: int = 1  # A
    data_start_row: int = 4
    data_end_row: int = 29

    @property
    def width(self) -> int:
        return self.end_col - self.start_col + 1

    @property
    def block_width(self) -> int:
        return self.width + self.spacer_cols


DEFAULT_LAYOUT = TableLayout()


def _iter_dates(start_date: date, end_date: date) -> list[date]:
    if start_date > end_date:
        raise ValueError("start_date must be <= end_date")

    days = (end_date - start_date).days
    return [start_date + timedelta(days=offset) for offset in range(days + 1)]


def _shift_formula_if_needed(value: object, source_cell: str, destination_cell: str) -> object:
    if not isinstance(value, str) or not value.startswith("="):
        return value
    return Translator(value, origin=source_cell).translate_formula(destination_cell)


def _copy_cell_value_and_style(
    ws: Worksheet,
    src_row: int,
    src_col: int,
    dst_row: int,
    dst_col: int,
) -> None:
    src_cell = ws.cell(src_row, src_col)
    dst_cell = ws.cell(dst_row, dst_col)

    src_addr = f"{get_column_letter(src_col)}{src_row}"
    dst_addr = f"{get_column_letter(dst_col)}{dst_row}"
    dst_cell.value = _shift_formula_if_needed(src_cell.value, src_addr, dst_addr)

    if src_cell.has_style:
        dst_cell._style = copy(src_cell._style)  # noqa: SLF001
    if src_cell.number_format is not None:
        dst_cell.number_format = copy(src_cell.number_format)
    if src_cell.font is not None:
        dst_cell.font = copy(src_cell.font)
    if src_cell.fill is not None:
        dst_cell.fill = copy(src_cell.fill)
    if src_cell.border is not None:
        dst_cell.border = copy(src_cell.border)
    if src_cell.alignment is not None:
        dst_cell.alignment = copy(src_cell.alignment)
    if src_cell.protection is not None:
        dst_cell.protection = copy(src_cell.protection)
    if src_cell.comment is not None:
        dst_cell.comment = copy(src_cell.comment)


def _copy_column_widths(ws: Worksheet, src_start_col: int, src_end_col: int, dst_start_col: int) -> None:
    for src_col in range(src_start_col, src_end_col + 1):
        offset = src_col - src_start_col
        dst_col = dst_start_col + offset
        src_key = get_column_letter(src_col)
        dst_key = get_column_letter(dst_col)
        ws.column_dimensions[dst_key].width = ws.column_dimensions[src_key].width


def _copy_row_heights(ws: Worksheet, start_row: int, end_row: int) -> None:
    for row_idx in range(start_row, end_row + 1):
        src_dim = ws.row_dimensions[row_idx]
        if src_dim.height is not None:
            ws.row_dimensions[row_idx].height = src_dim.height


def _copy_merged_ranges_for_block(
    ws: Worksheet,
    src_start_col: int,
    src_end_col: int,
    src_start_row: int,
    src_end_row: int,
    dst_start_col: int,
) -> None:
    col_offset = dst_start_col - src_start_col
    merged_ranges = list(ws.merged_cells.ranges)

    for merged in merged_ranges:
        min_col, min_row, max_col, max_row = merged.bounds
        if min_col < src_start_col or max_col > src_end_col:
            continue
        if min_row < src_start_row or max_row > src_end_row:
            continue

        ws.merge_cells(
            start_row=min_row,
            start_column=min_col + col_offset,
            end_row=max_row,
            end_column=max_col + col_offset,
        )


def _copy_table_block(ws: Worksheet, layout: TableLayout, table_index: int) -> None:
    if table_index == 0:
        return

    dst_start_col = layout.start_col + (table_index * layout.block_width)
    for row_idx in range(layout.start_row, layout.end_row + 1):
        for src_col in range(layout.start_col, layout.end_col + 1):
            offset = src_col - layout.start_col
            dst_col = dst_start_col + offset
            _copy_cell_value_and_style(ws, row_idx, src_col, row_idx, dst_col)

    _copy_column_widths(ws, layout.start_col, layout.end_col, dst_start_col)
    _copy_row_heights(ws, layout.start_row, layout.end_row)
    _copy_merged_ranges_for_block(
        ws,
        src_start_col=layout.start_col,
        src_end_col=layout.end_col,
        src_start_row=layout.start_row,
        src_end_row=layout.end_row,
        dst_start_col=dst_start_col,
    )


def _shift_cf_formula_cell_refs(formula: str, col_shift: int) -> str:
    def repl(match: re.Match[str]) -> str:
        p1, letters, p3, row = match.group(1), match.group(2), match.group(3), match.group(4)
        col_part = letters.replace("$", "")
        idx = column_index_from_string(col_part.upper())
        new_letters = get_column_letter(idx + col_shift)
        return f"{p1}{new_letters}{p3}{row}"

    return _CF_CELL_REF_RE.sub(repl, formula)


def _snapshot_conditional_formatting(template_ws: Worksheet) -> list:
    return list(template_ws.conditional_formatting)


def _append_shifted_conditional_formatting(
    ws: Worksheet,
    template_cfs: list,
    col_shift: int,
) -> None:
    """Duplicate template CF rules shifted horizontally (extra store table blocks)."""
    for cf in template_cfs:
        new_ranges: list[CellRange] = []
        for cr in cf.sqref:
            new_ranges.append(
                CellRange(
                    min_col=cr.min_col + col_shift,
                    min_row=cr.min_row,
                    max_col=cr.max_col + col_shift,
                    max_row=cr.max_row,
                )
            )
        new_sqref = MultiCellRange(new_ranges)
        sqref_str = str(new_sqref)

        for rule in cf.rules:
            new_rule = deepcopy(rule)
            if getattr(new_rule, "formula", None):
                new_rule.formula = [
                    _shift_cf_formula_cell_refs(f, col_shift) if isinstance(f, str) else f
                    for f in new_rule.formula
                ]
            ws.conditional_formatting.add(sqref_str, new_rule)


def _header_cells_for_table(layout: TableLayout, table_index: int) -> tuple[int, int, int]:
    start_col = layout.start_col + (table_index * layout.block_width)
    centech_col = start_col
    source_col = start_col + 2
    return start_col, centech_col, source_col


def _stamp_store_headers(
    ws: Worksheet,
    stores: list[str],
    source_label: str,
    layout: TableLayout,
) -> None:
    for idx, store in enumerate(stores):
        start_col, centech_col, source_col = _header_cells_for_table(layout, idx)
        ws.cell(1, start_col).value = f"Store {store}"
        ws.cell(2, centech_col).value = "CenTech"
        ws.cell(2, source_col).value = source_label


def _sheet_name_for_date(value: date, sheet_name_format: str) -> str:
    return value.strftime(sheet_name_format)


def build_template_workbook(
    *,
    template_path: Path,
    output_path: Path,
    start_date: date,
    end_date: date,
    stores: list[str],
    source_label: str,
    sheet_name_format: str = "%Y-%m-%d",
    layout: TableLayout = DEFAULT_LAYOUT,
) -> Path:
    if not stores:
        raise ValueError("At least one store is required.")

    workbook = load_workbook(template_path)
    template_sheet = workbook[workbook.sheetnames[0]]
    template_cf = _snapshot_conditional_formatting(template_sheet)

    target_dates = _iter_dates(start_date, end_date)
    for day in target_dates:
        new_sheet = workbook.copy_worksheet(template_sheet)
        new_sheet.title = _sheet_name_for_date(day, sheet_name_format)
        new_sheet.freeze_panes = template_sheet.freeze_panes
        # openpyxl's copy_worksheet does not copy conditional formatting; re-apply from the
        # template for every store table (col_shift=0 for the first block).
        for idx in range(len(stores)):
            if idx > 0:
                _copy_table_block(new_sheet, layout, idx)
            _append_shifted_conditional_formatting(
                new_sheet,
                template_cf,
                col_shift=idx * layout.block_width,
            )
        _stamp_store_headers(new_sheet, stores=stores, source_label=source_label, layout=layout)

    workbook.remove(template_sheet)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def category_rows_from_sheet(
    ws: Worksheet,
    *,
    layout: TableLayout = DEFAULT_LAYOUT,
) -> dict[str, int]:
    rows: dict[str, int] = {}
    for row_idx in range(layout.data_start_row, layout.data_end_row + 1):
        label = ws.cell(row_idx, layout.category_col).value
        if isinstance(label, str) and label.strip():
            rows[label.strip().lower()] = row_idx
    return rows


def apply_daily_store_values(
    ws: Worksheet,
    *,
    stores: list[str],
    daily_values: dict[str, dict[str, tuple[float, float, float, float]]],
    layout: TableLayout = DEFAULT_LAYOUT,
) -> None:
    category_to_row = category_rows_from_sheet(ws, layout=layout)

    for table_index, store in enumerate(stores):
        start_col = layout.start_col + (table_index * layout.block_width)
        centech_debit_col = start_col
        centech_credit_col = start_col + 1
        source_debit_col = start_col + 2
        source_credit_col = start_col + 3

        store_values = daily_values.get(str(store), {})
        for category_key, row_idx in category_to_row.items():
            values = store_values.get(category_key)
            if values is None:
                continue

            # Keep template formulas intact when present.
            centech_probe = ws.cell(row_idx, centech_debit_col).value
            source_probe = ws.cell(row_idx, source_debit_col).value
            if isinstance(centech_probe, str) and centech_probe.startswith("="):
                continue
            if isinstance(source_probe, str) and source_probe.startswith("="):
                continue

            c_debit, c_credit, s_debit, s_credit = values
            ws.cell(row_idx, centech_debit_col).value = c_debit
            ws.cell(row_idx, centech_credit_col).value = c_credit
            ws.cell(row_idx, source_debit_col).value = s_debit
            ws.cell(row_idx, source_credit_col).value = s_credit


def _cli_main() -> None:
    import argparse
    import sys

    from dateutil import parser as date_parser

    repo_root = Path(__file__).resolve().parents[3]
    sys.path.insert(0, str(repo_root))

    from financial.sales_export_comparison.rules import load_org_rule

    default_template = Path(__file__).resolve().parent.parent / "templates" / "Sales_Template.xlsx"
    default_rules = Path(__file__).resolve().parent.parent / "rules"

    def parse_date(raw: str):
        return date_parser.parse(raw, fuzzy=True).date()

    parser = argparse.ArgumentParser(
        description="Build only the Sales Comparison workbook shell (date tabs, store tables, headers).",
    )
    parser.add_argument("--start", type=str, required=True, help="Start date (e.g. 2026-03-01)")
    parser.add_argument("--end", type=str, required=True, help="End date (e.g. 2026-03-06)")
    parser.add_argument(
        "--output",
        type=str,
        required=True,
        help="Output .xlsx path (e.g. .../output/shell.xlsx)",
    )
    parser.add_argument("--template", type=str, default=str(default_template), help="Sales_Template.xlsx path")
    parser.add_argument(
        "--org",
        type=str,
        default=None,
        help="Load stores and sheet_date_format from rules/<org>.yaml",
    )
    parser.add_argument(
        "--stores",
        type=str,
        default=None,
        help="Comma-separated store numbers (required if --org omitted)",
    )
    parser.add_argument("--rules-dir", type=str, default=str(default_rules), help="Rules directory for --org")
    parser.add_argument(
        "--source-label",
        type=str,
        default="Client",
        help="Header label for client columns D/E",
    )
    parser.add_argument(
        "--sheet-date-format",
        type=str,
        default=None,
        help="strftime format for date tab names (default: %%b %%d or from org rules)",
    )
    args = parser.parse_args()

    start_date = parse_date(args.start)
    end_date = parse_date(args.end)
    if start_date > end_date:
        raise SystemExit("Start date must be <= end date.")

    if args.org:
        org_rule = load_org_rule(args.org, Path(args.rules_dir))
        stores = org_rule.stores
        sheet_fmt = args.sheet_date_format or org_rule.sheet_date_format
    else:
        if not args.stores:
            raise SystemExit("Provide --org or --stores.")
        stores = [s.strip() for s in args.stores.split(",") if s.strip()]
        if not stores:
            raise SystemExit("No stores in --stores.")
        sheet_fmt = args.sheet_date_format or "%b %d"

    out = build_template_workbook(
        template_path=Path(args.template),
        output_path=Path(args.output),
        start_date=start_date,
        end_date=end_date,
        stores=stores,
        source_label=args.source_label,
        sheet_name_format=sheet_fmt,
        layout=DEFAULT_LAYOUT,
    )
    print(f"Created: {out}")


if __name__ == "__main__":
    _cli_main()
