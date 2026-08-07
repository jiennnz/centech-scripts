from __future__ import annotations

from collections import defaultdict
from datetime import date, timedelta
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
WORKBOOK = ROOT / "financial" / "sales_export_comparison" / "runs" / "2026-06-01_2026-06-30" / "century" / "centech_vs_client" / "output" / "Sales_CenTech_vs_Client_2026-06-01_2026-06-30.xlsx"
AUDIT_OUT = ROOT / "financial" / "sales_export_comparison" / "runs" / "2026-06-01_2026-06-30" / "century" / "centech_vs_client" / "output" / "iscc_iscct_findings_audit.csv"
POS_DIR = ROOT / "pos_data"

ISCC = "In-Store Credit Card"
ISCCT = "In-Store Credit Card Tips"


def money(v: Any) -> float:
    return round(float(v or 0), 2)


def money_text(v: float) -> str:
    return f"${abs(v):,.2f}"


def scan_dates(target: str) -> list[str]:
    d = date.fromisoformat(target)
    start = d - timedelta(days=10)
    end = d + timedelta(days=30)
    out: list[str] = []
    cur = start
    while cur <= end:
        if (POS_DIR / cur.isoformat()).is_dir():
            out.append(cur.isoformat())
        cur += timedelta(days=1)
    return out


def load_store_id(target_date: str, store_number: str) -> str | None:
    path = POS_DIR / target_date / "Store.txt"
    if not path.exists():
        return None
    df = pd.read_csv(path, sep="|", dtype=str)
    hit = df[df["Store_Number"].astype(str).str.strip() == str(store_number)]
    if hit.empty:
        return None
    return str(hit["Store_ID"].iloc[0]).strip()


def load_store_payments(target_date: str, store_number: str) -> pd.DataFrame:
    store_id = load_store_id(target_date, store_number)
    if not store_id:
        return pd.DataFrame()

    pay_frames: list[pd.DataFrame] = []
    st_frames: list[pd.DataFrame] = []
    for folder in scan_dates(target_date):
        day_dir = POS_DIR / folder
        pay_path = day_dir / "Payment.txt"
        st_path = day_dir / "Sales_Ticket.txt"
        if not pay_path.exists() or not st_path.exists():
            continue
        pay = pd.read_csv(pay_path, sep="|", dtype=str)
        st = pd.read_csv(st_path, sep="|", dtype=str)
        pay["_pay_date"] = pd.to_datetime(pay["Payment_Date"], errors="coerce").dt.strftime("%Y-%m-%d")
        pay = pay[pay["_pay_date"] == target_date].copy()
        if pay.empty:
            continue
        pay["source_folder"] = folder
        ticket_dates = pay[["Ticket_Number", "_pay_date"]].dropna().drop_duplicates()
        st = st.merge(ticket_dates, on="Ticket_Number", how="inner")
        st = st[(st["_pay_date"] == target_date) & (st["Store_ID"].astype(str).str.strip() == store_id)].copy()
        if st.empty:
            continue
        st["sales_source_folder"] = folder
        pay_frames.append(pay)
        st_frames.append(st)

    if not pay_frames or not st_frames:
        return pd.DataFrame()

    sales = pd.concat(st_frames, ignore_index=True).drop_duplicates()
    payments = pd.concat(pay_frames, ignore_index=True)
    attrs = sales[
        ["Ticket_Number", "Status_ID", "Ticket_Type_ID", "Refund", "sales_source_folder"]
    ].drop_duplicates()
    df = payments.merge(attrs, on="Ticket_Number", how="inner")
    for col in ["Tendered_Amount", "Change", "Tip_Amount"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)
    for col in ["Payment_Type_ID", "Processing_Status_ID", "Transaction_ID", "Tip_Paid", "Ticket_Number", "Status_ID", "Ticket_Type_ID"]:
        df[col] = df[col].astype(str).str.strip()
    df["_tlen"] = df["Transaction_ID"].str.len()
    df["_base"] = df["Tendered_Amount"] - df["Change"]
    df["_iscc_value"] = df["_base"] + df["Tip_Amount"]
    return df


def get_discrepancies() -> list[dict[str, Any]]:
    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
    ws = wb["Discrepancies"]
    current_date: str | None = None
    rows: list[dict[str, Any]] = []
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        vals = list(row[:8])
        if vals[0] and isinstance(vals[0], str) and vals[0].startswith("Jun ") and "discrepancies" in vals[0]:
            current_date = f"2026-06-{vals[0].split()[1]}"
            continue
        if vals[1] not in {ISCC, ISCCT}:
            continue
        rows.append(
            {
                "excel_row": idx,
                "date": current_date,
                "store": str(vals[0]),
                "category": vals[1],
                "ct_debit": money(vals[2]),
                "ct_credit": money(vals[3]),
                "tqsr_debit": money(vals[4]),
                "tqsr_credit": money(vals[5]),
                "net_variance": money(vals[6]),
                "issue_type": vals[7],
            }
        )
    wb.close()
    return rows


def describe_crossdate(group: pd.DataFrame, amount: float, category: str, ct_higher: bool) -> tuple[str, str | None] | tuple[None, None]:
    if group.empty or "Payment_Type_ID" not in group.columns:
        return None, None
    t14 = group[group["Payment_Type_ID"] == "14"].copy()
    if t14.empty:
        return None, None
    status8 = t14[t14["Processing_Status_ID"] == "8"].copy()
    if not status8.empty:
        component = status8["Tip_Amount"] if category == ISCCT else status8["_iscc_value"]
        total = round(float(component.sum()), 2)
        if round(abs(total), 2) == round(abs(amount), 2):
            tickets = ", ".join(status8["Ticket_Number"].astype(str).drop_duplicates().tolist())
            side = "TQSR includes them while CenTech filters them out" if not ct_higher else "CenTech includes them while TQSR filters them out"
            label = "rows" if len(status8) > 1 else "row"
            return f"multiple provisional Processing_Status_ID 8 in-store card {label} ({tickets}); {side}.", None
    t14["component"] = t14["Tip_Amount"] if category == ISCCT else t14["_iscc_value"]
    candidates = t14[t14["component"].round(2).abs() == round(abs(amount), 2)]
    if candidates.empty and category == ISCC:
        candidates = t14[t14["_base"].round(2).abs() == round(abs(amount), 2)]
    if candidates.empty:
        return None, None
    for ticket, g in candidates.groupby("Ticket_Number"):
        all_rows = t14[t14["Ticket_Number"] == ticket]
        folders = list(dict.fromkeys(all_rows["source_folder"].astype(str)))
        statuses = list(dict.fromkeys(all_rows["Processing_Status_ID"].astype(str)))
        if len(folders) > 1:
            original = folders[0]
            updated = folders[-1]
            side = "TQSR includes an earlier POS copy while CenTech uses the later update" if not ct_higher else "CenTech includes the later POS update while TQSR is still on the earlier copy"
            return f"ticket {ticket} cross-date update: original folder {original}, updated/finalized folder {updated}; {side}.", str(ticket)
        if "8" in statuses:
            side = "TQSR includes it while CenTech filters it out" if not ct_higher else "CenTech includes it while TQSR filters it out"
            return f"ticket {ticket} filter issue: provisional Processing_Status_ID 8 in-store card row; {side}.", str(ticket)
        if any(s not in {"4", "2"} for s in statuses):
            side = "TQSR includes it while CenTech filters it out" if not ct_higher else "CenTech includes it while TQSR filters it out"
            return f"ticket {ticket} filter issue: Processing_Status_ID {','.join(statuses)} card row; {side}.", str(ticket)
    return None, None


def describe_gift_card(group: pd.DataFrame, amount: float, ct_higher: bool) -> tuple[str, str | None] | tuple[None, None]:
    if group.empty or "Payment_Type_ID" not in group.columns:
        return None, None
    t14_gc = group[(group["Payment_Type_ID"] == "14") & (group["Status_ID"] == "8")].copy()
    if t14_gc.empty:
        return None, None
    grouped = t14_gc.groupby("Ticket_Number", as_index=False)["_base"].sum()
    exact = grouped[grouped["_base"].round(2).abs() == round(abs(amount), 2)]
    if exact.empty:
        # Allow multiple gift-card-sold tickets to explain a single ISCC-only delta.
        total = round(float(grouped["_base"].sum()), 2)
        if round(abs(total), 2) != round(abs(amount), 2):
            return None, None
        tickets = ", ".join(grouped["Ticket_Number"].astype(str).tolist())
    else:
        tickets = ", ".join(exact["Ticket_Number"].astype(str).tolist())
    side = "CenTech counts it in ISCC; TQSR filters it out" if ct_higher else "TQSR counts it in ISCC; CenTech filters it out"
    label = "tickets" if "," in tickets else "ticket"
    return f"gift-card-sold {label} {tickets} paid by in-store credit card; {side}.", tickets


def describe_ticket(group: pd.DataFrame, ticket: str, amount: float, category: str, ct_higher: bool) -> str:
    t14 = group[(group["Payment_Type_ID"] == "14") & (group["Ticket_Number"] == str(ticket))].copy()
    folders = list(dict.fromkeys(t14["source_folder"].astype(str))) if not t14.empty else []
    statuses = list(dict.fromkeys(t14["Processing_Status_ID"].astype(str))) if not t14.empty else []
    if len(folders) > 1:
        side = "TQSR includes an earlier POS copy while CenTech uses the later update" if not ct_higher else "CenTech includes the later POS update while TQSR is still on the earlier copy"
        return f"ticket {ticket} cross-date update: original folder {folders[0]}, updated/finalized folder {folders[-1]}; {side}."
    if "8" in statuses:
        side = "TQSR includes it while CenTech filters it out" if not ct_higher else "CenTech includes it while TQSR filters it out"
        return f"ticket {ticket} filter issue: provisional Processing_Status_ID 8 in-store card row; {side}."
    side = "CenTech includes this ticket and TQSR excludes it" if ct_higher else "TQSR includes this ticket and CenTech excludes it"
    return f"ticket {ticket} filter/date issue; {side}."


def fallback(group: pd.DataFrame, amount: float, category: str, ct_higher: bool) -> tuple[str, str | None]:
    if group.empty or "Payment_Type_ID" not in group.columns:
        return "No matching POS payment rows found in the cross-date scan window; likely export-side classification/date issue.", None
    t14 = group[group["Payment_Type_ID"] == "14"].copy()
    if t14.empty:
        return "No matching type-14 POS payment rows found in the cross-date scan window; likely export-side classification/date issue.", None
    if category == ISCCT:
        cand = t14[t14["Tip_Amount"].round(2).abs() == round(abs(amount), 2)]
    else:
        cand = t14[t14["_iscc_value"].round(2).abs() == round(abs(amount), 2)]
        if cand.empty:
            cand = t14[t14["_base"].round(2).abs() == round(abs(amount), 2)]
    if not cand.empty:
        ticket = str(cand.iloc[0]["Ticket_Number"])
        status = str(cand.iloc[0]["Processing_Status_ID"])
        folder = str(cand.iloc[0]["source_folder"])
        side = "CenTech includes this row and TQSR excludes it" if ct_higher else "TQSR includes this row and CenTech excludes it"
        return f"ticket {ticket} filter/date issue: type-14 row in folder {folder} with Processing_Status_ID {status}; {side}.", ticket
    side = "CenTech is higher than TQSR" if ct_higher else "TQSR is higher than CenTech"
    return f"{side}; no single POS card row exactly matches the variance, review aggregate ISCC/ISCCT filter/date handling for this store-day.", None


def build_findings(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    cache: dict[tuple[str, str], pd.DataFrame] = {}
    out: list[dict[str, Any]] = []
    row_ticket: dict[tuple[str, str, str], str] = {}
    for item in rows:
        key = (item["date"], item["store"])
        if key not in cache:
            cache[key] = load_store_payments(*key)
        group = cache[key]
        category = item["category"]
        amount = item["net_variance"]
        ct_higher = amount > 0
        finding = None
        ticket = None
        if category == ISCC:
            finding, ticket = describe_gift_card(group, amount, ct_higher)
        if not finding:
            finding, ticket = describe_crossdate(group, amount, category, ct_higher)
        if not finding:
            finding, ticket = fallback(group, amount, category, ct_higher)
        if ticket:
            row_ticket[(item["date"], item["store"], item["category"])] = ticket
        item["finding"] = finding
        out.append(item)

    # If the tip row identified the exact ticket, use it for the paired ISCC debit
    # row when the debit row only reached an aggregate fallback.
    for item in out:
        if item["category"] != ISCC:
            continue
        if "no single POS card row exactly matches" not in item["finding"]:
            continue
        ticket = row_ticket.get((item["date"], item["store"], ISCCT))
        if not ticket:
            continue
        key = (item["date"], item["store"])
        group = cache[key]
        item["finding"] = describe_ticket(group, ticket, item["net_variance"], item["category"], item["net_variance"] > 0)
    return out


def write_workbook(findings: list[dict[str, Any]]) -> None:
    wb = load_workbook(WORKBOOK)
    ws = wb["Discrepancies"]
    row_to_finding = {x["excel_row"]: x["finding"] for x in findings}
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 8).value == "Issue Type":
            ws.cell(r, 9).value = "Finding"
    for r, finding in row_to_finding.items():
        ws.cell(r, 9).value = finding
    wb.save(WORKBOOK)
    wb.close()


def main() -> None:
    rows = get_discrepancies()
    findings = build_findings(rows)
    pd.DataFrame(findings).to_csv(AUDIT_OUT, index=False)
    write_workbook(findings)
    print(f"Annotated {len(findings)} ISCC/ISCCT discrepancies")
    print(f"Workbook: {WORKBOOK}")
    print(f"Audit CSV: {AUDIT_OUT}")


if __name__ == "__main__":
    main()
